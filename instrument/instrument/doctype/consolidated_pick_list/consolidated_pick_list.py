# Copyright (c) 2022, instrument and contributors
# For license information, please see license.txt

# import frappe
# from frappe.model.document import Document

# class ConsolidatedPickList(Document):
# 	pass


# Copyright (c) 2021, instrument and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from erpnext.manufacturing.doctype.bom.bom import get_bom_items_as_dict
from erpnext.stock.doctype.item.item import get_item_defaults
# from erpnext.stock.doctype.batch.batch import get_batch_no
from frappe import _ , msgprint
from frappe.query_builder.functions import Truncate
from frappe.utils import cint, flt, get_link_to_form,comma_and

import json
from datetime import datetime
from frappe.utils import nowdate,nowtime, today, flt
from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
from erpnext.manufacturing.doctype.bom.bom import get_children, validate_bom_no
from frappe.utils import nowdate,nowtime, today, flt
from datetime import timedelta,date
import datetime
import calendar
import json
import time

class ConsolidatedPickList(Document):
	@frappe.whitelist()
	def get_fg_work_orders(self):
		filters={"work_order":self.work_order, "planned_start_date":self.planned_start_date, "planned_end_date":self.planned_end_date, "expected_delivery_date":self.expected_delivery_date,"production_planning_with_lead_time":self.production_plan}
		# self.work_order_table = ''
		fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
		todays_date = datetime.date.today()
		fg_item_groups = [item.get('item_group') for item in fg_item_groups]
		joined_fg_list = "', '".join(fg_item_groups)
		if fg_item_groups:
			if self.purpose == "Material Transfer for Manufacture By FG":
				fg_work_orders = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.material_transferred_for_manufacturing) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where i.item_group in ('{0}') and wo.planned_start_date >= '{1}' and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and (wo.qty - wo.material_transferred_for_manufacturing) > 0 and  {2}""".format(joined_fg_list,todays_date, wo_filter_condition(filters)),as_dict=1,debug=1)
				
				if fg_work_orders:
					for row in fg_work_orders:
						doc = frappe.get_doc("Work Order",row.get("name"))
						ohs = get_current_stock()
						bom_childs = []
						bom_child_list = get_child_boms(doc.bom_no,bom_childs)
						bom_child_list.append({'bom' : doc.bom_no})
						
						# find all the bom in descending order (bom level)
						# print("=============",bom_child_list)
						final_bom_list = get_all_boms_in_order(bom_child_list)
						
						final_bom_list = [item.get('bom') for item in final_bom_list if item.get('bom')]
						final_qty_dict = dict()
						for bom in final_bom_list:
							qty_will_be_produced_list = []
							bom_item = frappe.db.get_value("BOM",{'name':bom},'item')
							qty_will_be_produced = 0
							
							bom_items = get_raw_bom_data(bom)
							
							if bom_items:
								for item in bom_items:
									if item.item_code in ohs:
										qty_will_be_produced = qty_will_be_produced + (ohs.get(item.item_code)/item.qty)
									if item.item_code in final_qty_dict:
										qty_will_be_produced = qty_will_be_produced + final_qty_dict.get(item.item_code)
									qty_will_be_produced_list.append(qty_will_be_produced)
								if qty_will_be_produced_list:
									final_qty_dict[bom_item] = min(qty_will_be_produced_list)
								else:
									final_qty_dict[bom_item] = 0

						row['qty_will_be_produced'] =final_qty_dict.get(doc.production_item) if doc.production_item in final_qty_dict else 0 
						bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
						self.append('work_order_table',{
							'work_order':row.get('name'),
							'pending_qty':row.get('pending_qty'),
							'qty_can_be_pulledmanufactured' :row.get('qty_will_be_produced'),
							'total_qty_of_finish_good_on_work_order':row.get('qty'),
							'qty_of_finished_goods_already_completed':row.get('produced_qty'),
							'item_to_manufacture':doc.production_item,
							'item_name':doc.item_name,
							'bom_no':doc.bom_no,
							'bom_level':bom_level,
							'planned_start_date':doc.planned_start_date,
							'planned_end_date':doc.planned_end_date,
							'actual_start_date':doc.actual_start_date,
							'actual_end_date':doc.actual_end_date

							})
			elif self.purpose == "Material Transfer for Manufacture By Work Order":
				fg_work_orders = frappe.db.sql("""SELECT distinct wo.name,(wo.qty - wo.material_transferred_for_manufacturing) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and (wo.qty - wo.material_transferred_for_manufacturing) > 0 and  {0} order by wo.bom_level,wo.name""".format( wo_filter_condition(filters)),as_dict=1,debug=1)

				
				if fg_work_orders:
					for row in fg_work_orders:
						doc = frappe.get_doc("Work Order",row.get("name"))
						ohs = get_current_stock()
	
						bom_childs = []
						bom_child_list = get_child_boms(doc.bom_no,bom_childs)
						bom_child_list.append({'bom' : doc.bom_no})
						
						# find all the bom in descending order (bom level)
						# print("=============",bom_child_list)
						final_bom_list = get_all_boms_in_order(bom_child_list)
						
						final_bom_list = [item.get('bom') for item in final_bom_list if item.get('bom')]
						final_qty_dict = dict()
						for bom in final_bom_list:
							qty_will_be_produced_list = []
							bom_item = frappe.db.get_value("BOM",{'name':bom},'item')
							qty_will_be_produced = 0
							
							bom_items = get_raw_bom_data(bom)
							
							if bom_items:
								for item in bom_items:
									if item.item_code in ohs:
										# print("--------------------------------------------777----------", ohs.get(item.item_code))
										# print("--------------------------------------------777----------", item.item_code)
										qty_will_be_produced = qty_will_be_produced + (ohs.get(item.item_code)/item.qty)
									if item.item_code in final_qty_dict:
										qty_will_be_produced = qty_will_be_produced + final_qty_dict.get(item.item_code)
									qty_will_be_produced_list.append(qty_will_be_produced)
								if qty_will_be_produced_list:
									final_qty_dict[bom_item] = min(qty_will_be_produced_list)
								else:
									final_qty_dict[bom_item] = 0

						row['qty_will_be_produced'] =final_qty_dict.get(doc.production_item) if doc.production_item in final_qty_dict else 0 
						bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
						self.append('work_order_table',{
							'work_order':row.get('name'),
							'pending_qty':row.get('pending_qty'),
							'qty_can_be_pulledmanufactured' :row.get('qty_will_be_produced'),
							'total_qty_of_finish_good_on_work_order':row.get('qty'),
							'qty_of_finished_goods_already_completed':row.get('produced_qty'),
							'item_to_manufacture':doc.production_item,
							'item_name':doc.item_name,
							'bom_no':doc.bom_no,
							'bom_level':bom_level,
							'planned_start_date':doc.planned_start_date,
							'planned_end_date':doc.planned_end_date,
							'actual_start_date':doc.actual_start_date,
							'actual_end_date':doc.actual_end_date

							})
			elif self.purpose == "Manufacture By FG":
				filters={"work_order":self.work_order, "planned_start_date":self.planned_start_date, "planned_end_date":self.planned_end_date, "expected_delivery_date":self.expected_delivery_date,"production_planning_with_lead_time":self.production_plan}
				fg_work_orders = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where i.item_group in ('{0}') and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0  and (wo.qty - wo.produced_qty) > 0 and {1}""".format(joined_fg_list, wo_filter_condition(filters)),as_dict=1,debug=1)
				fg_work_orders_2 = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where i.item_group in ('{0}') and wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 and wo.produced_qty < wo.qty and (wo.qty - wo.produced_qty) > 0  and {1}""".format(joined_fg_list, wo_filter_condition(filters)),as_dict=1,debug=1)
				if fg_work_orders_2:
					for item in fg_work_orders_2:
						fg_work_orders.append(item)
				if fg_work_orders:
					for row in fg_work_orders:
						doc = frappe.get_doc("Work Order",row.get("name"))
						if doc.skip_transfer == 1 and doc.from_wip_warehouse == 1:
							ohs = get_current_stock_for_manufacture()
						elif doc.skip_transfer == 1:
							ohs = get_current_stock()
						else:
							ohs = get_current_stock_for_manufacture()
						bom_childs = []
						bom_child_list = get_child_boms(doc.bom_no,bom_childs)
						bom_child_list.append({'bom' : doc.bom_no})
						
						# find all the bom in descending order (bom level)
						final_bom_list = get_all_boms_in_order(bom_child_list)
						final_bom_list = [item.get('bom') for item in final_bom_list]
						final_qty_dict = dict()
						for bom in final_bom_list:
							qty_will_be_produced_list = []
							bom_item = frappe.db.get_value("BOM",{'name':bom},'item')
							qty_will_be_produced = 0
							bom_items = get_raw_bom_data(bom)
							if bom_items:
								for item in bom_items:
									if item.item_code in ohs:
										qty_will_be_produced = qty_will_be_produced + (ohs.get(item.item_code)/item.qty)
									if item.item_code in final_qty_dict:
										qty_will_be_produced = qty_will_be_produced + final_qty_dict.get(item.item_code)
									qty_will_be_produced_list.append(qty_will_be_produced)
								if qty_will_be_produced_list:
									final_qty_dict[bom_item] = min(qty_will_be_produced_list)
								else:
									final_qty_dict[bom_item] = 0
						
						row['qty_will_be_produced'] =final_qty_dict.get(doc.production_item) if doc.production_item in final_qty_dict else 0 
						bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
						self.append('work_order_table',{
							'work_order':row.get('name'),
							'pending_qty':row.get('pending_qty'),
							'qty_can_be_pulledmanufactured' :row.get('qty_will_be_produced'),
							'total_qty_of_finish_good_on_work_order':row.get('qty'),
							'qty_of_finished_goods_already_completed':row.get('produced_qty'),
							'item_to_manufacture':doc.production_item,
							'item_name':doc.item_name,
							'bom_no':doc.bom_no,
							'bom_level':bom_level,
							'planned_start_date':doc.planned_start_date,
							'planned_end_date':doc.planned_end_date,
							'actual_start_date':doc.actual_start_date,
							'actual_end_date':doc.actual_end_date
							})
			else:
				filters={"work_order":self.work_order, "planned_start_date":self.planned_start_date, "planned_end_date":self.planned_end_date, "expected_delivery_date":self.expected_delivery_date,"production_planning_with_lead_time":self.production_plan}
				fg_work_orders = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where wo.status in ('Not Started','In Process') and wo.skip_transfer = 0  and (wo.qty - wo.produced_qty) > 0 and {0}""".format(wo_filter_condition(filters)),as_dict=1,debug=1)
				fg_work_orders_2 = frappe.db.sql("""SELECT wo.name,(wo.qty - wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 and wo.produced_qty < wo.qty and (wo.qty - wo.produced_qty) > 0  and {0}""".format(wo_filter_condition(filters)),as_dict=1,debug=1)
				if fg_work_orders_2:
					for item in fg_work_orders_2:
						fg_work_orders.append(item)
				if fg_work_orders:
					for row in fg_work_orders:
						doc = frappe.get_doc("Work Order",row.get("name"))
						if doc.skip_transfer == 1 and doc.from_wip_warehouse == 1:
							ohs = get_current_stock_for_manufacture()
						elif doc.skip_transfer == 1:
							ohs = get_current_stock()
						else:
							ohs = get_current_stock_for_manufacture()
						bom_childs = []
						bom_child_list = get_child_boms(doc.bom_no,bom_childs)
						bom_child_list.append({'bom' : doc.bom_no})
						
						# find all the bom in descending order (bom level)
						final_bom_list = get_all_boms_in_order(bom_child_list)
						final_bom_list = [item.get('bom') for item in final_bom_list]
						final_qty_dict = dict()
						for bom in final_bom_list:
							qty_will_be_produced_list = []
							bom_item = frappe.db.get_value("BOM",{'name':bom},'item')
							qty_will_be_produced = 0
							bom_items = get_raw_bom_data(bom)
							if bom_items:
								for item in bom_items:
									if item.item_code in ohs:
										qty_will_be_produced = qty_will_be_produced + (ohs.get(item.item_code)/item.qty)
									if item.item_code in final_qty_dict:
										qty_will_be_produced = qty_will_be_produced + final_qty_dict.get(item.item_code)
									qty_will_be_produced_list.append(qty_will_be_produced)
								if qty_will_be_produced_list:
									final_qty_dict[bom_item] = min(qty_will_be_produced_list)
								else:
									final_qty_dict[bom_item] = 0
						
						row['qty_will_be_produced'] =final_qty_dict.get(doc.production_item) if doc.production_item in final_qty_dict else 0 
						bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
						self.append('work_order_table',{
							'work_order':row.get('name'),
							'pending_qty':row.get('pending_qty'),
							'qty_can_be_pulledmanufactured' :row.get('qty_will_be_produced'),
							'total_qty_of_finish_good_on_work_order':row.get('qty'),
							'qty_of_finished_goods_already_completed':row.get('produced_qty'),
							'item_to_manufacture':doc.production_item,
							'item_name':doc.item_name,
							'bom_no':doc.bom_no,
							'bom_level':bom_level,
							'planned_start_date':doc.planned_start_date,
							'planned_end_date':doc.planned_end_date,
							'actual_start_date':doc.actual_start_date,
							'actual_end_date':doc.actual_end_date
							})
		else:
			frappe.throw("Please Enter Item Groups for FG in Rushabh Settings")
	
	@frappe.whitelist()
	def get_sub_assembly_items(self, manufacturing_type=None):
		if self.purpose in ['Material Transfer for Manufacture By Work Order','Manufacture By Work Order']:
			self.sub_assembly_items = []
			for row in self.work_order_table:
				bom = frappe.db.get_value("Work Order",{'name':row.work_order},'bom_no')
				doc = frappe.get_doc("Work Order",row.work_order)
				planned_start_date = frappe.db.get_value("Work Order",{'name':row.work_order},'planned_start_date')
				bom_data = []
				get_sub_assembly_items(bom, bom_data, row.pending_qty)
				bom_qty_dict = {item.production_item : item.stock_qty for item in bom_data}
				sub_assembly_items = [item.production_item for item in bom_data]
				
				final_subassembly_items = []
				if self.consider_current_stock == 1:
					wip_warehouse = frappe.db.get_value("Work Order",{'name':row.work_order},'wip_warehouse')
					current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse = '{0}' and actual_qty > 0 group by item_code """.format(wip_warehouse),as_dict=1)
					ohs_dict = {item.item_code : item.qty for item in current_stock}
					for col in sub_assembly_items:
						if col not in ohs_dict:
							final_subassembly_items.append(col)
						elif (ohs_dict.get(col) < bom_qty_dict.get(col)):
							final_subassembly_items.append(col)
					# self.get_work_orders_for_subassembly(final_subassembly_items,planned_start_date,row.work_order,bom_qty_dict)
				else:
					print("--------------111")
					# self.get_work_orders_for_subassembly(sub_assembly_items,planned_start_date,row.work_order,bom_qty_dict)
				
				job_cards = frappe.db.sql("""SELECT jb.name,jb.status from `tabJob Card` jb where jb.work_order = '{0}'""".format(row.work_order),as_dict=1)
				if job_cards:
					job_cards = [i.name for i in job_cards]
					job_cards = [get_link_to_form('Job Card', p) for p in job_cards]
				
				work_order_status = frappe.db.get_value("Work Order",{'name':row.work_order},'status')
				bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
				if self.workstation:
					job_cards = frappe.db.sql("""SELECT jb.name,jb.status from `tabJob Card` jb where jb.work_order = '{0}' and jb.workstation = '{1}'""".format(row.work_order,self.workstation),as_dict=1)
					if job_cards != []:
						self.append('work_orders',{
							'work_order':row.work_order,
							'qty_of_finished_goods_to_pull':row.pending_qty,
							'total_qty_to_of_finished_goods_on_work_order':row.total_qty_of_finish_good_on_work_order,
							'qty_of_finished_goods_already_completed':row.qty_of_finished_goods_already_completed,
							'qty_of_finished_goods':row.qty_can_be_pulledmanufactured,
							'work_order_status':work_order_status,
							'job_cards': str(job_cards),
							'item_to_manufacture':doc.production_item,
							'item_name':doc.item_name,
							'bom_no':doc.bom_no,
							'bom_level':bom_level,
							'planned_start_date':doc.planned_start_date,
							'planned_end_date':doc.planned_end_date,
							'actual_start_date':doc.actual_start_date,
							'actual_end_date':doc.actual_end_date
							})
				else:
					self.append('work_orders',{
							'work_order':row.work_order,
							'qty_of_finished_goods_to_pull':row.pending_qty,
							'total_qty_to_of_finished_goods_on_work_order':row.total_qty_of_finish_good_on_work_order,
							'qty_of_finished_goods_already_completed':row.qty_of_finished_goods_already_completed,
							'qty_of_finished_goods':row.qty_can_be_pulledmanufactured,
							'work_order_status':work_order_status,
							'job_cards': str(job_cards),
							'item_to_manufacture':doc.production_item,
							'item_name':doc.item_name,
							'bom_no':doc.bom_no,
							'bom_level':bom_level,
							'planned_start_date':doc.planned_start_date,
							'planned_end_date':doc.planned_end_date,
							'actual_start_date':doc.actual_start_date,
							'actual_end_date':doc.actual_end_date
							})
		if self.purpose in ['Material Transfer for Manufacture By FG','Manufacture By FG']:
			self.sub_assembly_items = []
			for row in self.work_order_table:
				bom = frappe.db.get_value("Work Order",{'name':row.work_order},'bom_no')
				doc = frappe.get_doc("Work Order",row.work_order)
				planned_start_date = frappe.db.get_value("Work Order",{'name':row.work_order},'planned_start_date')
				bom_data = []
				get_sub_assembly_items(bom, bom_data, row.pending_qty)
				bom_qty_dict = {item.production_item : item.stock_qty for item in bom_data}
				sub_assembly_items = [item.production_item for item in bom_data]
				
				final_subassembly_items = []
				if self.consider_current_stock == 1:
					wip_warehouse = frappe.db.get_value("Work Order",{'name':row.work_order},'wip_warehouse')
					current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse = '{0}' and actual_qty > 0 group by item_code """.format(wip_warehouse),as_dict=1)
					ohs_dict = {item.item_code : item.qty for item in current_stock}
					for col in sub_assembly_items:
						if col not in ohs_dict:
							final_subassembly_items.append(col)
						elif (ohs_dict.get(col) < bom_qty_dict.get(col)):
							final_subassembly_items.append(col)
					self.get_work_orders_for_subassembly(final_subassembly_items,planned_start_date,row.work_order,bom_qty_dict)
				else:
					
					self.get_work_orders_for_subassembly(sub_assembly_items,planned_start_date,row.work_order,bom_qty_dict)
				
				job_cards = frappe.db.sql("""SELECT jb.name,jb.status from `tabJob Card` jb where jb.work_order = '{0}'""".format(row.work_order),as_dict=1)
				if job_cards:
					job_cards = [i.name for i in job_cards]
					job_cards = [get_link_to_form('Job Card', p) for p in job_cards]
				
				work_order_status = frappe.db.get_value("Work Order",{'name':row.work_order},'status')
				bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
				self.append('work_orders',{
					'work_order':row.work_order,
					'qty_of_finished_goods_to_pull':row.pending_qty,
					'total_qty_to_of_finished_goods_on_work_order':row.total_qty_of_finish_good_on_work_order,
					'qty_of_finished_goods_already_completed':row.qty_of_finished_goods_already_completed,
					'qty_of_finished_goods':row.qty_can_be_pulledmanufactured,
					'work_order_status':work_order_status,
					'job_cards': str(job_cards),
					'item_to_manufacture':doc.production_item,
					'item_name':doc.item_name,
					'bom_no':doc.bom_no,
					'bom_level':bom_level,
					'planned_start_date':doc.planned_start_date,
					'planned_end_date':doc.planned_end_date,
					'actual_start_date':doc.actual_start_date,
					'actual_end_date':doc.actual_end_date
					})


	def get_work_orders_for_subassembly(self,sub_assembly_items,planned_start_date,work_order,bom_qty_dict):
		todays_date = datetime.date.today()
		sub_assembly_items = "', '".join(sub_assembly_items)
		sales_order = frappe.db.get_value("Work Order",work_order,'so_reference')
		mr_reference = frappe.db.get_value("Work Order",work_order,'mr_reference')
		parent_item_code = frappe.db.get_value("Work Order",work_order,'parent_item_code')
		if self.purpose in ['Material Transfer for Manufacture By FG','Material Transfer for Manufacture By Work Order']:
			work_orders = frappe.db.sql("""SELECT production_item,name,(qty-produced_qty) as pending_qty ,qty,produced_qty from `tabWork Order` where production_item in ('{0}') and planned_start_date <= '{1}' and docstatus =1 and status in ('Not Started','In Process') and (so_reference = '{2}' or mr_reference = '{3}') order by name desc""".format(sub_assembly_items,planned_start_date.date(),sales_order,mr_reference),as_dict=1,debug=1)
		else:
			# work_orders =frappe.db.sql("""SELECT wo.production_item,wo.name,(wo.qty-wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo where wo.production_item in ('{0}') and wo.planned_start_date <= '{1}' and wo.docstatus =1 and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.produced_qty < wo.material_transferred_for_manufacturing and wo.material_transferred_for_manufacturing > 0 and (so_reference = '{2}' or mr_reference = '{3}')""".format(sub_assembly_items,planned_start_date.date(),sales_order,mr_reference),as_dict=1,debug=1)
			work_orders =frappe.db.sql("""SELECT wo.production_item,wo.name,(wo.qty-wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo where wo.production_item in ('{0}') and wo.planned_start_date <= '{1}' and wo.docstatus =1 and wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.produced_qty < wo.material_transferred_for_manufacturing and (so_reference = '{2}' or mr_reference = '{3}')""".format(sub_assembly_items,planned_start_date.date(),sales_order,mr_reference),as_dict=1,debug=1)
			work_order_1 =frappe.db.sql("""SELECT wo.production_item,wo.name,(wo.qty-wo.produced_qty) as pending_qty ,wo.qty,wo.produced_qty from `tabWork Order` wo where wo.production_item in ('{0}') and wo.planned_start_date <= '{1}' and wo.docstatus =1 and wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 and wo.produced_qty < wo.qty and (so_reference = '{2}' or mr_reference = '{3}')""".format(sub_assembly_items,planned_start_date.date(),sales_order,mr_reference),as_dict=1,debug=1)
			if work_order_1:
				for item in work_order_1:
					work_orders.append(item)
		if work_orders:
			for row in work_orders:
				doc = frappe.get_doc("Work Order",row.get("name"))
				if doc.skip_transfer == 1 and doc.from_wip_warehouse == 1:
					ohs = get_current_stock_for_manufacture()
				elif self.purpose == "Manufacture":
					ohs = get_current_stock_for_manufacture()
				elif doc.skip_transfer == 1:
					ohs = get_current_stock()
				else:
					ohs = get_current_stock()
				qty_will_be_produced_list = []
				bom_childs = []
				bom_child_list = get_child_boms(doc.bom_no,bom_childs)
				bom_child_list.append({'bom' : doc.bom_no})
				
				# find all the bom in descending order (bom level)
				final_bom_list = get_all_boms_in_order(bom_child_list)
				final_bom_list = [item.get('bom') for item in final_bom_list]
				final_qty_dict = dict()
				for bom in final_bom_list:
					qty_will_be_produced_list = []
					bom_item = frappe.db.get_value("BOM",{'name':bom},'item')
					qty_will_be_produced = 0
					bom_items = get_raw_bom_data(bom)
					if bom_items:
						for item in bom_items:
							if item.item_code in ohs:
								qty_will_be_produced = qty_will_be_produced + (ohs.get(item.item_code)*item.qty)
							if item.item_code in final_qty_dict:
								qty_will_be_produced = qty_will_be_produced + final_qty_dict.get(item.item_code)
							qty_will_be_produced_list.append(qty_will_be_produced)
						if qty_will_be_produced_list:
							final_qty_dict[bom_item] = min(qty_will_be_produced_list)
						else:
							final_qty_dict[bom_item] = 0
				# for item in doc.required_items:
				# 	if item.item_code in ohs:
				# 		if item.required_qty <= ohs.get(item.item_code):
				# 			percent_stock = 100
				# 			qty_will_be_produced = item.required_qty
				# 			qty_will_be_produced_list.append(qty_will_be_produced)
				# 		elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
				# 			percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
				# 			qty_will_be_produced = (percent_stock/100*item.required_qty)
				# 			qty_will_be_produced_list.append(qty_will_be_produced)
				# 		else : 
				# 			percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
				# 			qty_will_be_produced = 0
				# 			qty_will_be_produced_list.append(qty_will_be_produced)
				# 	else:
				# 		qty_will_be_produced = 0
				# 		qty_will_be_produced_list.append(qty_will_be_produced)
				row['qty_will_be_produced'] =final_qty_dict.get(doc.production_item) if doc.production_item in final_qty_dict else 0 
				job_cards = frappe.db.sql("""SELECT jb.name,jb.status from `tabJob Card` jb where jb.work_order = '{0}'""".format(row.get('name')),as_dict=1)
				if job_cards:
					job_cards = [i.name for i in job_cards]
					job_cards = [get_link_to_form('Job Card', p) for p in job_cards]
				work_order_status = frappe.db.get_value("Work Order",{'name':row.get('name')},'status')
				bom_level = frappe.db.get_value("BOM",{'name':doc.bom_no},'bom_level')
				self.append('work_orders',{
					'work_order':row.get('name'),
					'qty_of_finished_goods_to_pull':bom_qty_dict.get(row.get('production_item')),
					'total_qty_to_of_finished_goods_on_work_order':row.get('qty'),
					'qty_of_finished_goods_already_completed':row.get('produced_qty'),
					'qty_of_finished_goods':row.get('qty_will_be_produced'),
					'job_cards':str(job_cards),
					'work_order_status':work_order_status,
					'item_to_manufacture':doc.production_item,
					'item_name':doc.item_name,
					'bom_no':doc.bom_no,
					'bom_level':bom_level,
					'planned_start_date':doc.planned_start_date,
					'planned_end_date':doc.planned_end_date,
					'actual_start_date':doc.actual_start_date,
					'actual_end_date':doc.actual_end_date
					})

	def validate(self):
		if self.docstatus == 0:
			final_item_list = []
			work_order_list = [row.get('work_order') for row in self.work_order_table]
			dup = {x for x in work_order_list if work_order_list.count(x) > 1}
			if dup :
				frappe.throw("There are duplicates work orders selected in Work Order Table <b>{0}</b>".format(dup))
			for row in self.work_orders:
				final_item_list.append({
				'work_order' :row.get("work_order"),
				'total_qty_to_of_finished_goods_on_work_order':row.get("total_qty_to_of_finished_goods_on_work_order"),
				'qty_of_finished_goods_to_pull':row.get("qty_of_finished_goods_to_pull"),
				'qty_of_finished_goods_already_completed':row.get("qty_of_finished_goods_already_completed"),
				'qty_of_finished_goods':row.get("qty_of_finished_goods"),
				'job_cards':row.get('job_cards'),
				'work_order_status':row.get('work_order_status'),
				'item_to_manufacture':row.get('item_to_manufacture'),
				'item_name':row.get('item_name'),
				'bom_no':row.get('bom_no'),
				'bom_level':row.get('bom_level'),
				'planned_start_date':row.get('planned_start_date'),
				'planned_end_date':row.get('planned_end_date'),
				'actual_start_date':row.get('actual_start_date'),
				'actual_end_date':row.get('actual_end_date')
				})
			final_data = sorted(final_item_list,key = lambda x:x["bom_level"])
			self.work_orders = []
			count = 1
			for col in final_data:
				col.update({'idx' : count})
				self.append("work_orders",col)
				count = count + 1

			for wo in work_order_list:
				wo_doc = frappe.get_doc("Work Order",wo)
				if wo_doc:
					ref = frappe.db.get_value("Consolidated Pick List Reference",{'parent':wo_doc.name,'consolidated_pick_list':self.name})
					if ref:
						frappe.db.set_value("Consolidated Pick List Reference",{'parent':wo_doc.name,'consolidated_pick_list':self.name},'status',self.status)
						frappe.db.commit()
					else:
						# frappe.db.set_value("Consolidated Pick List Reference",{'parent':wo_doc.name},'consolidated_pick_list',self.name)
						# frappe.db.set_value("Consolidated Pick List Reference",{'parent':wo_doc.name},'status',self.status)
						# frappe.db.commit()
						wo_doc.append('consolidated_pick_list_reference',{
							'consolidated_pick_list':self.name,
							'status':self.status
							})
						wo_doc.save()
			self.update_status()
	@frappe.whitelist()
	def get_work_order_items(self):
		final_raw_item_list = []
		
		final_data = dict()
		if self.work_orders: 
			for item in self.work_orders:
				item_list = []
				bom_no = frappe.db.get_value("Work Order",{'name':item.work_order},'bom_no')
				wo_doc = frappe.get_doc("Work Order",item.work_order)
				if bom_no:
					# Get all the required raw materials with required qty according to qty of qty_of_finished_goods
					raw_materials = get_raw_material(bom_no,self.company,item.qty_of_finished_goods_to_pull,item.work_order)
					final_raw_item_list.append(raw_materials)
					i_list = [item for item in raw_materials]
					for i in i_list:
						item_list.append(i)
					final_item_list= list(set(item_list))
					# Get all the avaialble locations for required items
					if self.purpose in ['Material Transfer for Manufacture By FG','Material Transfer for Manufacture By Work Order']:
						item_locations_dict = get_item_locations(self,final_item_list,self.company)
					if self.purpose in ['Manufacture By FG','Manufacture By Work Order']:
						if wo_doc.skip_transfer == 1 and wo_doc.from_wip_warehouse == 0:
							item_locations_dict = get_item_locations(self,final_item_list,self.company)
						else:
							item_locations_dict = get_item_locations_for_manufacture(self,final_item_list,self.company)
					# Allocate qty from wip warehouse
					# consider wip stock for the item and consume first wip stock,if wip stock allocated then do not consider it for next row 
					wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
					
					wip_qty = get_wip_stock()
					remaining_wip_qty = dict()
		
					# Manipulate in order to show in table
					for row in raw_materials:
						engineering_revision = frappe.db.get_value("Work Order Item",{'parent':item.work_order,'item_code':row},'engineering_revision')
						if self.purpose in ['Material Transfer for Manufacture By FG','Material Transfer for Manufacture By Work Order']:
							if row in wip_qty and row not in remaining_wip_qty:
								qty = abs(wip_qty.get(row) - raw_materials.get(row).get('qty')) if wip_qty.get(row) < raw_materials.get(row).get('qty') else 0
								rem_qty = abs(wip_qty.get(row) - raw_materials.get(row).get('qty')) if wip_qty.get(row) > raw_materials.get(row).get('qty') else 0
								raw_materials.get(row).update({'qty':qty})
								raw_materials.get(row)['wip_stock'] =rem_qty
							elif row in remaining_wip_qty:
								raw_materials.get(row).update({'qty' : 0 if remaining_wip_qty.get(row) >= raw_materials.get(row).get('qty') else 0 })
								
								remaining_wip_qty.update({row:abs(remaining_wip_qty.get(row)-raw_materials.get(row).get('qty'))})
								raw_materials.get(row)['wip_stock'] = remaining_wip_qty.get(row)
			

						if row in item_locations_dict:
							col = item_locations_dict.get(row)
							for i in col:
								# Fetch batch  FIFO basis
								# batch_no = get_batch_no(row, i.get("warehouse"),raw_materials.get(row).get("qty"), throw=False)
								transferred_qty = frappe.db.get_value("Work Order Item",{'parent':item.work_order,'item_code':row},'transferred_qty')
								# i['required_qty'] = (flt(raw_materials.get(row).get("qty"))-flt(transferred_qty)) 
								i['required_qty'] = flt(raw_materials.get(row).get("qty"))
								i['picked_qty'] = 1
								i['stock_uom'] = raw_materials.get(row).get('stock_uom')
								i['engineering_revision'] = engineering_revision
								i['wip_stock'] = flt(raw_materials.get(row).get("wip_stock"))
								# i['batch_no'] = batch_no
								# Prefill the picked qty
								# if batch_no:
								# 	batch_qty = frappe.db.get_value("Batch",{'name':batch_no},'batch_qty')
								# 	if batch_qty >= (flt(raw_materials.get(row).get("qty"))-flt(transferred_qty)) :
								# 		i['picked_qty'] = (flt(raw_materials.get(row).get("qty"))-flt(transferred_qty)) 
								# 	else:
								# 		i['picked_qty'] = batch_qty
					final_data[item.work_order] = item_locations_dict
			# add items in item_locations
			work_order_list = [item.work_order for item in self.work_orders]
			f_item_list = []
			f_item_data = []
			for work_order in work_order_list:
				
				for row in final_data:
					if row == work_order:
						final_row = final_data.get(row)
						for i in final_row:
							item_data = get_item_defaults(i, self.company)
							j = final_row.get(i)
							for d in j:
								if (d.get("required_qty")) > 0:
									f_item_list.append(i)
									self.append("work_order_pick_list_item",{
										"item_code": i,
										"item_name": item_data.get("item_name"),
										"warehouse":d.get("warehouse"),
										"required_qty":d.get("required_qty"),
										"stock_qty": d.get("qty"),
										"work_order" : row,
										"picked_qty" : d.get('picked_qty'),
										"uom" : d.get('stock_uom'),
										"stock_uom":d.get('stock_uom'),
										"description" :item_data.get('description'),
										"item_group" : item_data.get("item_group"),
										"engineering_revision" : d.get("engineering_revision"),
										"wip_stock" : d.get("wip_stock")
										# "batch_no" : d.get('batch_no')
									})
									f_item_data.append({
										"item_code": i,
										"item_name": item_data.get("item_name"),
										"warehouse":d.get("warehouse"),
										"required_qty":d.get("required_qty"),
										"stock_qty": d.get("qty"),
										"work_order" : row,
										"picked_qty" : d.get('picked_qty'),
										"uom" : d.get('stock_uom'),
										"stock_uom":d.get('stock_uom'),
										"description" :item_data.get('description'),
										"item_group" : item_data.get("item_group"),
										"engineering_revision" : d.get("engineering_revision"),
										"wip_stock" : d.get("wip_stock")
										# "batch_no" : d.get('batch_no')
									})
			for work_order in work_order_list:
				doc = frappe.get_doc("Work Order", work_order)
				for unstock_rec in doc.unstock_items_table:
					self.append("unstock_items_table",{
										"item_code": unstock_rec.item_code,
										"item_name": unstock_rec.item_name,
										"description":unstock_rec.description,
										"qty":unstock_rec.qty,
									})

			x = self.batch_assignment_fifo(f_item_list)

			final_x_data = []
			for f_rec in f_item_data:
				for x_rec in x:
					if f_rec.get('item_code') == x_rec.get('item_code') and f_rec.get("warehouse") == x_rec.get('warehouse'):
						f_rec['batch_no'] = x_rec.get('batch_no')
						final_x_data.append(f_rec)

			return final_x_data
		# self.save()
	# def batch_assignment_fifo(self):
	# 	item_list = [item.item_code for item in self.work_order_pick_list_item]
	# 	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
		
	# 	batch_data = frappe.db.sql("""SELECT b.name,b.item,`tabStock Ledger Entry`.warehouse, sum(`tabStock Ledger Entry`.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` ignore index (item_code, warehouse) on (b.name = `tabStock Ledger Entry`.batch_no ) where `tabStock Ledger Entry`.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) and `tabStock Ledger Entry`.warehouse != '{1}'  group by batch_id order by b.expiry_date ASC, b.creation ASC""".format(tuple(item_list),wip_warehouse),as_dict=1,debug=1)
	# 	allocated_item_list = []
	# 	allocated_item_dict = dict()
	# 	for row in self.work_order_pick_list_item:
	# 		if row.item_code not in allocated_item_dict:
	# 			for col in batch_data:	
	# 				if row.item_code == col.item and row.warehouse == col.warehouse :
						
	# 					if col.qty >= row.required_qty:
	# 						row.batch_no = col.name
	# 						row.picked_qty = row.required_qty
	# 						allocated_item_list.append(row.item_code)
	# 						allocated_item_dict.update({row.item_code:row.work_order})
	# 						if frappe.get_cached_value('Item', row.item_code, 'has_serial_no') == 1:
	# 							serial_nos = get_serial_no_batchwise(row.item_code,col.name,col.warehouse,row.required_qty)
	# 							row.serial_no = serial_nos 
	# 						break
	# 					elif col.qty < row.required_qty:
	# 						row.batch_no = col.name
	# 						row.picked_qty = col.qty
	# 						allocated_item_list.append(row.item_code)
	# 						# allocated_item_dict.update({row.item_code:row.work_order})
	# 						if frappe.get_cached_value('Item', row.item_code, 'has_serial_no') == 1:
	# 							serial_nos = get_serial_no_batchwise(row.item_code,col.name,col.warehouse,row.required_qty)
	# 							row.serial_no = serial_nos
	# 						break

	def batch_assignment_fifo(self,f_item_list):
		item_list = [item.item_code for item in self.work_order_pick_list_item]
		if f_item_list:
			for i in f_item_list:
				item_list.append(i)
		joined_item_list = "', '".join(item_list)
		wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
		work_list = []
		for row in self.work_order_pick_list_item:
			work_list.append({"item_code":row.item_code, "uom":row.uom, "uom_conversion_factor":row.uom_conversion_factor, "stock_uom":row.stock_uom, "serial_nos":row.serial_no, "warehouse":row.warehouse, "required_qty":row.required_qty, "work_order":row.work_order, "stock_qty":0, "picked_qty":0,"wip_stock":row.wip_stock})
		
		if item_list:
			if self.purpose in ['Material Transfer for Manufacture By FG','Material Transfer for Manufacture By Work Order']:
				# batch_data = frappe.db.sql("""SELECT b.name,b.item, sle.warehouse,sum(sle.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` sle on (b.name = sle.batch_no ) where sle.item_code in ('{0}')  and IFNULL(b.expiry_date, '2200-01-01') > %(today)s and sle.warehouse != '{1}' and b.disabled = 0 and sle.is_cancelled=0 and b.batch_qty > 0 group by sle.warehouse,b.name order by  b.creation ASC""".format(joined_item_list,wip_warehouse),{'today':today()},as_dict=1,debug=1)
				# batch_data = frappe.db.sql("""SELECT b.name,b.item, sle.warehouse,sum(sle.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` sle on (b.name = sle.batch_no ) where sle.item_code in ('{0}')  and IFNULL(b.expiry_date, '2200-01-01') > %(today)s and sle.warehouse != '{1}' and b.disabled = 0 and sle.is_cancelled=0 group by b.name,sle.warehouse having qty > 0 order by  b.creation ASC""".format(joined_item_list,wip_warehouse),{'today':today()},as_dict=1,debug=1)
				# batch_data  = frappe.db.sql(
				# 		"""
				# 		SELECT
				# 			sle.`warehouse` as warehouse,
				# 			sle.`batch_no` as name,
				# 			SUM(sle.`actual_qty`) AS `qty`,
				# 			sle.item_code as item
				# 		FROM
				# 			`tabStock Ledger Entry` sle, `tabBatch` batch
				# 		WHERE
				# 			sle.batch_no = batch.name
				# 			and sle.`item_code` in ('{0}')
				# 			and batch.disabled = 0
				# 			and sle.is_cancelled=0
				# 			and sle.warehouse != '{1}'
				# 		GROUP BY
				# 			sle.`batch_no`
				# 		HAVING `qty` > 0
				# 		ORDER BY IFNULL(batch.`expiry_date`, '2200-01-01'), batch.`creation`
				# 	""".format(joined_item_list,wip_warehouse),
				# 		as_dict=1,debug=1
				# 	)
				batch_data  = frappe.db.sql(
						"""
						SELECT
							sle.`warehouse` as warehouse,
							sle.`serial_and_batch_bundle` as serial_and_batch_bundle,
							b.`name` as name , 
							SUM(sle.`actual_qty`) AS `qty`,
							sle.`item_code` as item
						FROM
							`tabStock Ledger Entry` sle, `tabSerial and Batch Entry` sbe ,`tabBatch` b
						WHERE
							sle.`serial_and_batch_bundle` = sbe.`parent` and sbe.`batch_no` = b.`name`
							and sle.`item_code` in ('{0}')
							and b.`disabled` = 0
							and sle.`is_cancelled`=0
							and sle.`warehouse` != '{1}'
						GROUP BY
							sbe.`batch_no`
						HAVING `qty` > 0
						ORDER BY IFNULL(b.`expiry_date`, '2200-01-01'), b.`creation`
					""".format(joined_item_list,wip_warehouse),
						as_dict=1,debug=1
					)
			elif self.purpose in ['Manufacture By FG','Manufacture By Work Order']:
				if self.change_batches_as_per_previous_consolidated_pick_list == 1:
					# for row in work_list:
					work_order_list = [row.get("work_order") for row in work_list]
					work_order_list ="', '".join(work_order_list)
					having_clause = "having sum(sle.actual_qty) > 0"
					# correct_batch_nos = frappe.db.sql("""SELECT sle.batch_no, round(sum(sle.actual_qty),2) from `tabStock Ledger Entry` sle
					# 	INNER JOIN `tabBatch` batch on sle.batch_no = batch.name where
					# 	batch.disabled = 0
					# 	and sle.is_cancelled = 0
					# 	and sle.warehouse = '{0}'
					# 	and batch.docstatus < 2
					# group by batch_no having sum(sle.actual_qty) > 0
					# order by batch.expiry_date, sle.batch_no desc""".format(wip_warehouse),as_dict=1)
					correct_batch_nos = frappe.db.sql("""SELECT batch.`name`, round(sum(sle.`actual_qty`),2) from `tabStock Ledger Entry` sle,`tabSerial and Batch Entry` sbe, `tabBatch` batch where sle.`serial_and_batch_bundle` = sbe.`parent` and  sbe.`batch_no` = batch.`name` and
						batch.`disabled` = 0
						and sle.`is_cancelled` = 0
						and sle.`warehouse` = '{0}'
						and batch.`docstatus` < 2
					group by batch.`name` having sum(sle.`actual_qty`) > 0
					order by batch.`expiry_date`, sbe.`batch_no` desc""".format(wip_warehouse),as_dict=1,debug=1)
					correct_batch_nos =  [row.get('name') for row in correct_batch_nos]
					correct_batch_nos ="', '".join(correct_batch_nos)
					batch_data = frappe.db.sql("""SELECT co.name as consolidated_pick_list,wo.batch_no as name,wo.work_order,wo.item_code as item,wo.picked_qty as qty from `tabWork Order Pick List Item` wo join `tabConsolidated Pick List` co on co.name = wo.parent where wo.work_order in ('{0}') and wo.batch_no in ('{1}') and co.purpose = 'Material Transfer for Manufacture By FG' and co.docstatus=1 order by wo.work_order,wo.item_code""".format(work_order_list,correct_batch_nos),as_dict=1,debug=0)
					# batch_nos = [row.get('name') for row in batch_data]
					# batch_nos ="', '".join(batch_nos)
					
					for row in batch_data:
						row['warehouse'] = wip_warehouse
						batch_doc = frappe.get_doc("Batch",row.get('name'))
						if batch_doc.batch_qty <= 0:
							batch_data.remove(row)
						
				else:
					# batch_data = frappe.db.sql("""SELECT b.name,b.item, sle.warehouse, sum(sle.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` sle on (b.name = sle.batch_no ) where sle.item_code in ('{0}')  and IFNULL(b.expiry_date, '2200-01-01') > %(today)s and sle.warehouse = '{1}' and b.disabled = 0 and sle.is_cancelled=0 group by b.name,sle.warehouse HAVING qty > 0 order by  b.creation ASC""".format(joined_item_list,wip_warehouse),{'today':today()},as_dict=1,debug=1)
					# batch_data  = frappe.db.sql(
					# 	"""
					# 	SELECT
					# 		sle.`warehouse` as warehouse,
					# 		sle.`batch_no` as name,
					# 		SUM(sle.`actual_qty`) AS `qty`,
					# 		sle.item_code as item
					# 	FROM
					# 		`tabStock Ledger Entry` sle, `tabBatch` batch
					# 	WHERE
					# 		sle.batch_no = batch.name
					# 		and sle.`item_code` in ('{0}')
					# 		and batch.disabled = 0
					# 		and sle.is_cancelled=0
					# 		and sle.warehouse = '{1}'
					# 	GROUP BY
					# 		sle.`batch_no`
					# 	HAVING `qty` > 0
					# 	ORDER BY IFNULL(batch.`expiry_date`, '2200-01-01'), batch.`creation`
					# """.format(joined_item_list,wip_warehouse),
					# 	as_dict=1,debug=1
					# )
					batch_data  = frappe.db.sql(
						"""
						SELECT
							sle.`warehouse` as warehouse,
							sle.`serial_and_batch_bundle` as serial_and_batch_bundle,
							b.`name` as name , 
							SUM(sle.`actual_qty`) AS `qty`,
							sle.`item_code` as item
						FROM
							`tabStock Ledger Entry` sle, `tabSerial and Batch Entry` sbe ,`tabBatch` b
						WHERE
							sle.`serial_and_batch_bundle` = sbe.`parent` and sbe.`batch_no` = b.`name`
							and sle.`item_code` in ('{0}')
							and b.`disabled` = 0
							and sle.`is_cancelled`=0
							and sle.`warehouse` = '{1}'
						GROUP BY
							sbe.`batch_no`
						HAVING `qty` > 0
						ORDER BY IFNULL(b.`expiry_date`, '2200-01-01'), b.`creation`
					""".format(joined_item_list,wip_warehouse),
						as_dict=1,debug=1
					)
			# for row in batch_data:
			# 	calculate_remaining_batch_qty(row)
			new_list=[]
			allocated_item_dict = dict()
			allocated_war_item_dict = dict()
		
			for row in work_list:
				wo_item = row.get("item_code")+row.get("work_order")
				war_item = row.get("work_order")+row.get("item_code")+row.get("warehouse")
				row.update({'actual_required_qty':row.get('required_qty')})

				for col in batch_data:
					if row.get("item_code")==col.get("item") and row.get("warehouse")==col.get("warehouse") and wo_item not in allocated_item_dict:
						item_qty=sum([r.get("picked_qty") for r in new_list if new_list and row.get("item_code")==r.get("item_code") and r.get("item_code")+r.get("work_order")==wo_item])           
						rem_reqd_qty = abs(item_qty-row.get("actual_required_qty"))
						row.update({"required_qty":rem_reqd_qty})
						if col.get("qty")>= row.get("required_qty"):
							batch_qty = col.get("qty") - row.get("required_qty") 
							new_list.append({
								"item_code":row.get("item_code"), 
			                    "warehouse":row.get("warehouse"), 
			                    "required_qty":row.get("required_qty"),
			                    "work_order":row.get("work_order"),
			                    "stock_qty":col.get("qty"),
			                    "picked_qty":row.get("required_qty"),
			                    "batch_no": col.get("name"),
								"uom":row.get("uom"), 
								"uom_conversion_factor":row.get("uom_conversion_factor"), 
								"stock_uom":row.get("stock_uom"),
								"wip_stock" : row.get('wip_stock'),
								"actual_required_qty":row.get('actual_required_qty')
			                })
							col.update({"qty":batch_qty})
							allocated_item_dict.update({wo_item:wo_item})
						else:
							reqd_qty = row.get("required_qty")-col.get("qty")
							if col.get("qty") > 0:
								new_list.append({
									"item_code":row.get("item_code"), 
									"warehouse":row.get("warehouse"), 
									"required_qty":row.get("required_qty"),
									"work_order":row.get("work_order"),
			                        "stock_qty":col.get("qty"),
			                        "picked_qty":col.get("qty"),
			                        "batch_no": col.get("name"),
			                        "uom":row.get("uom"), 
									"uom_conversion_factor":row.get("uom_conversion_factor"), 
									"stock_uom":row.get("stock_uom"),
									"wip_stock" : row.get('wip_stock'),
									"actual_required_qty":row.get('actual_required_qty')
			                    })
			      #           elif war_item not in allocated_war_item_dict:
			      #               new_list.append({
			      #                   "item_code":row.get("item_code"), 
			      #                   "warehouse":row.get("warehouse"), 
			      #                   "required_qty":row.get("required_qty"),
			      #                   "work_order":row.get("work_order"),
			      #                   "stock_qty": 0,
			      #                   "picked_qty": 0,
			      #                   "batch_no": "",
			      #                   "uom":row.get("uom"), 
									# "uom_conversion_factor":row.get("uom_conversion_factor"), 
									# "stock_uom":row.get("stock_uom")
			      #               })
			      #               allocated_war_item_dict.update({war_item:war_item})
							col.update({"qty":0})		
			return new_list
		else:
			frappe.msgprint("No Items Found For Pick")
	@frappe.whitelist()
	def remove_work_orders(self):
		pass
	#Pick Sales Order Pick Qty From Batch Basis On FIFO
	@frappe.whitelist()
	def get_sales_order_items(self):
		final_raw_item_list = []
		final_data = []
		main_item_code = ""
		if self.pick_list_sales_order_table:
			for item in self.pick_list_sales_order_table:
				# Get all the avaialble locations for required items
				if self.purpose == 'Sales Order Fulfillment':
					item_locations_dict = get_so_and_po_item_locations(self, item.item, self.company, item.qty, item.sales_order, main_item_code,item.sales_order_item)
					final_data.append(item_locations_dict)
			
			# add items in item_locations
			# sales_order_list = [item.sales_order for item in self.pick_list_sales_order_table]
			# # for sales_order in sales_order_list:
			for item in final_data:
				for row in item:
					item_data = get_item_defaults(row, self.company)
					for itm in item.get(row):
						if flt(itm.get("required_qty")) > 0:
							self.append("sales_order_pick_list_item",{
								"item_code": row,
								"item_name": item_data.get("item_name"),
								"warehouse":itm.get("warehouse"),
								"required_qty":itm.get("required_qty"),
								"stock_qty": itm.get("qty"),
								"sales_order" : itm.get("order"),
								"picked_qty" : itm.get('picked_qty') or 1,
								"uom" : itm.get('stock_uom'),
								"stock_uom":itm.get('stock_uom'),
								"description":item_data.get('description'),
								"item_group" : item_data.get("item_group"),
								"sales_order_item":itm.get("sales_order_item")
							})
			 
			return self.so_batch_assignment_fifo()


	def so_batch_assignment_fifo(self):
		item_list = [item.item_code for item in self.sales_order_pick_list_item]
		wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
		item_list =  '(' + ','.join("'{}'".format(i) for i in item_list) + ')' if item_list else ()

		# batch_data = frappe.db.sql("""SELECT b.name,b.item, `tabStock Ledger Entry`.warehouse, sum(`tabStock Ledger Entry`.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` ignore index (item_code, warehouse) on (b.name = `tabStock Ledger Entry`.batch_no ) where `tabStock Ledger Entry`.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) and `tabStock Ledger Entry`.warehouse != '{1}'  group by batch_id order by b.expiry_date ASC, b.creation ASC""".format(item_list, wip_warehouse),as_dict=1,debug=1)

		batch_data = frappe.db.sql("""SELECT b.name,b.item,sle.warehouse, sum(sle.actual_qty) as qty from `tabBatch` b ,`tabSerial and Batch Entry` sbe, `tabStock Ledger Entry` sle where sle.serial_and_batch_bundle = sbe.parent and sbe.batch_no = b.name and sle.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) group by b.name order by b.expiry_date ASC, b.creation ASC""".format(item_list, wip_warehouse),as_dict=1,debug=1)
		
		for row in batch_data:
			calculate_remaining_so_batch_qty(row)

		so_list = []
		for row in self.sales_order_pick_list_item:
			so_list.append({"item_code":row.item_code, "uom":row.uom, "uom_conversion_factor":row.uom_conversion_factor, "stock_uom":row.stock_uom, "serial_nos":row.serial_no, "warehouse":row.warehouse, "required_qty":row.required_qty, "sales_order":row.sales_order, "stock_qty":0, "picked_qty":1,"sales_order_item":row.sales_order_item})
			# print("===============so_list",so_list)
		
		new_list=[]
		allocated_item_dict = dict()
		allocated_war_item_dict = dict()
		for row in so_list:
			so_item = row.get("item_code")+row.get("sales_order")
			# war_item = row.get("sales_order")+row.get("item_code")+row.get("warehouse")
			for col in batch_data:
				if row.get("item_code")==col.get("item") and row.get("warehouse")==col.get("warehouse") and so_item not in allocated_item_dict:
					item_qty=sum([r.get("picked_qty") for r in new_list if new_list and row.get("item_code")==r.get("item_code") and r.get("item_code")+r.get("sales_order")==so_item])            
					rem_reqd_qty = abs(item_qty-row.get("required_qty"))
					row.update({"required_qty":rem_reqd_qty})

					serial_no = ""
					if frappe.get_cached_value("Item", row.get("item_code"), 'has_serial_no') == 1:
						serial_nos = get_serial_no_batchwise(row.get("item_code"),col.get("name"), row.get("warehouse"), row.get("required_qty"))
						serial_no = serial_nos

					if col.get("qty")>= row.get("required_qty"):
						batch_qty = col.get("qty") - row.get("required_qty") 
						new_list.append({
							"item_code":row.get("item_code"), 
							"warehouse":row.get("warehouse"), 
							"required_qty":row.get("required_qty"),
							"sales_order":row.get("sales_order"),
							"stock_qty":col.get("qty"),
							"picked_qty":row.get("required_qty"),
							"batch_no": col.get("name"),
							"uom":row.get("uom"), 
							"uom_conversion_factor":row.get("uom_conversion_factor"), 
							"stock_uom":row.get("stock_uom"),
							"serial_no": serial_no,
							"sales_order_item":row.get("sales_order_item")
		                })
						col.update({"qty":batch_qty})
						allocated_item_dict.update({so_item:so_item})
					else:
						reqd_qty = row.get("required_qty")-col.get("qty")
						if col.get("qty") > 0:
							new_list.append({
								"item_code":row.get("item_code"), 
								"warehouse":row.get("warehouse"), 
								"required_qty":row.get("required_qty"),
								"sales_order":row.get("sales_order"),
								"stock_qty":col.get("qty"),
								"picked_qty":col.get("qty"),
								"batch_no": col.get("name"),
								"uom":row.get("uom"), 
								"uom_conversion_factor":row.get("uom_conversion_factor"), 
								"stock_uom":row.get("stock_uom"),
								"serial_no": serial_no,
								"sales_order_item":row.get("sales_order_item")
							})
						col.update({"qty":0})		
		return new_list

	def on_submit(self):
		allocated_batch_data = frappe.db.sql("""SELECT b.item_code,b.picked_qty,b.batch_no from `tabWork Order Pick List Item` b join `tabConsolidated Pick List` wo on wo.name = b.parent""",as_dict=1,debug=0)
		allocated_batch_data_dict = {item.batch_no:item.picked_qty for item in allocated_batch_data}
		allocated_batch_list = []
		for row in self.work_order_pick_list_item:
			if row.batch_no in allocated_batch_data_dict:
				allocated_batch_list.append(row.idx)
		# if len(allocated_batch_list) > 0:
		# 	frappe.throw("Kindly Review batches for rows {0}".format(allocated_batch_list))
		self.update_status()

	def update_status(self):
		status = []
		if self.purpose == "Material Transfer for Manufacture By FG":
			for row in self.work_orders:
				check_stock_entry = frappe.db.get_value("Stock Entry",{'work_order':row.work_order,'consolidated_pick_list':self.name,'purpose':"Material Transfer for Manufacture"},['name','docstatus'],as_dict=1)
				# print("===========check_stock_entry",check_stock_entry)
				if check_stock_entry:
					row.stock_entry = check_stock_entry.get('name')
					row.stock_entry_status = "Submitted" if check_stock_entry.get('docstatus') == 1 else 'Draft'
		elif self.purpose == "Manufacture By FG":
			for row in self.work_orders:
				check_stock_entry = frappe.db.get_value("Stock Entry",{'work_order':row.work_order,'consolidated_pick_list':self.name,'purpose':"Manufacture"},['name','docstatus'],as_dict=1)
				# print("===========check_stock_entry",check_stock_entry)
				if check_stock_entry:
					row.stock_entry = check_stock_entry.get('name')
					row.stock_entry_status = "Submitted" if check_stock_entry.get('docstatus') == 1 else 'Draft'


				# if row.qty_of_finished_goods>0 and not status:
				# 	row.stock_entry_status="Not Created"
				# 	status.append("Not Created")
				# 	frappe.db.commit()

	@frappe.whitelist()
	def get_fg_sales_orders(self):
		filters={"delivery_start_date":self.delivery_start_date, "delivery_end_date":self.delivery_end_date, "customer":self.customer, "customer_po_number":self.customer_po_number, "sales_order":self.sales_order}
		self.sales_order_table = []
		self.purpose == "Sales Order Fulfillment"
		fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
		todays_date = datetime.date.today()
		fg_item_groups = [item.get('item_group') for item in fg_item_groups]
		if len(fg_item_groups) == 0:
			frappe.throw("Please Enter Item Groups for FG in Rushabh Settings")
		joined_fg_list = "', '".join(fg_item_groups)
		if fg_item_groups:
			if self.purpose == "Sales Order Fulfillment":
				fg_sales_orders = frappe.db.sql("""SELECT so.name,sum(i.qty - i.delivered_qty) as pending_qty , i.item_code, i.qty, i.produced_qty ,i.name as sales_order_item from `tabSales Order` so join `tabSales Order Item` i on i.parent = so.name where i.item_group in ('{0}') and so.delivery_date >= '{1}' and so.status not in ('Completed','Cancel') and (i.qty - i.delivered_qty) > 0 and {2} group by so.name order by so.name""".format(joined_fg_list,todays_date, so_filter_condition(filters)),as_dict=1,debug=1)

				# sales_order = frappe.db.sql("""SELECT so.name, sum(i.qty - i.delivered_qty) as pending_qty ,i.name as sales_order_item from `tabSales Order` so join `tabSales Order Item` i on i.parent = so.name where i.item_group in ('{0}') and so.delivery_date >= '{1}' and so.status not in ('Completed','Cancel') and {2} group by so.name order by so.name """.format(joined_fg_list,todays_date, so_filter_condition(filters)), as_dict=True, debug=1)

				for row in fg_sales_orders:
					self.append('sales_order_table',{
							'sales_order':row.get('name'),
							'pending_qty':row.get('pending_qty')
						})

	#get Sales Orders
	@frappe.whitelist()
	def get_sales_orders(self):
		qty_prod = {}
		ohs = get_current_stock()
		for itm in self.sales_order_table:
			fg_sales_orders = frappe.db.sql("""SELECT so.name,(i.qty - i.delivered_qty) as pending_qty , i.item_code, i.qty, i.produced_qty ,i.name as sales_order_item from `tabSales Order` so join `tabSales Order Item` i on i.parent = so.name where so.name='{0}' order by so.name""".format(itm.sales_order), as_dict=1)

			for row in fg_sales_orders:
				if row.get("item_code") in ohs:
					if row.get("pending_qty") <= ohs.get(row.get("item_code")):
						qty_will_be_delivered = row.get("pending_qty")
						remaining_qty = flt(ohs.get(row.get("item_code"))) - flt(row.get("pending_qty"))
						ohs.update({row.get("item_code"):remaining_qty})
						row['qty_can_be_pulled'] = qty_will_be_delivered
					elif row.get("pending_qty") > ohs.get(row.get("item_code")) and ohs.get(row.get("item_code")) >= 0: 
						qty_will_be_delivered = ohs.get(row.get("item_code"))
						ohs.update({row.get("item_code"):0})
						row['qty_can_be_pulled'] = qty_will_be_delivered
				else : 
					qty_will_be_delivered = 0
					row['qty_can_be_pulled'] = qty_will_be_delivered
				self.append('pick_list_sales_order_table',{
					'sales_order':row.get('name'),
					'item':row.get('item_code'),
					'qty':row.get('pending_qty'),
					'qty_can_be_pulled' :row.get('qty_can_be_pulled'),
					'sales_order_item': row.get('sales_order_item')
				})
		return True


	#Get FG Purchase Orders
	@frappe.whitelist()
	def get_fg_purchase_orders(self):
		filters={"start_date":self.start_date, "end_date":self.end_date, "supplier":self.supplier, "purchase_order": self.purchase_order}
		self.purchase_order_table = []
		self.purpose == "Material Transfer for Subcontracted Goods"
		fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
		todays_date = datetime.date.today()
		fg_item_groups = [item.get('item_group') for item in fg_item_groups]
		joined_fg_list = "', '".join(fg_item_groups)
		if fg_item_groups:
			if self.purpose == "Material Transfer for Subcontracted Goods":
				fg_purchase_orders = frappe.db.sql("""SELECT po.name, sum(i.qty) as pending_qty , i.item_code ,i.name as purchase_order_item from `tabPurchase Order` po join `tabPurchase Order Item` i on i.parent = po.name where i.item_group in ('{0}') and po.schedule_date >= '{1}' and po.docstatus=1 and po.status not in ('Completed','Cancel') and po.is_subcontracted=1 and {2} group by po.name order by po.name""".format(joined_fg_list,todays_date, po_filter_condition(filters)),as_dict=1,debug=1)

				# purchase_order = frappe.db.sql("""SELECT po.name, sum(i.qty) as pending_qty, i.item_code ,i.name as purchase_order_item  from `tabPurchase Order` po join `tabPurchase Order Item` i on i.parent = po.name where i.item_group in ('{0}') and po.schedule_date >= '{1}' and po.docstatus=1 and po.status not in ('Completed','Cancel') and po.is_subcontracted='Yes' and {2} group by po.name""".format(joined_fg_list,todays_date, po_filter_condition(filters)),as_dict=1,debug=0)

				if fg_purchase_orders:
					for row in fg_purchase_orders:
						self.append('purchase_order_table',{
							'purchase_order':row.get('name'),
							'pending_qty':row.get('pending_qty')
						})
				else:
					frappe.msgprint("No Purchase Orders Found")
		else:
			frappe.throw("Please Enter Item Groups for FG in Rushabh Settings")

	#get Purchase Orders
	@frappe.whitelist()
	def get_purchase_orders(self):
		qty_prod = {}
		ohs = get_current_stock()
		for itm in self.purchase_order_table:
			fg_purchase_orders = frappe.db.sql("""SELECT po.name, i.qty as pending_qty , i.item_code ,i.fg_item,i.name as purchase_order_item from `tabPurchase Order` po join `tabPurchase Order Item` i on i.parent = po.name where po.name='{0}' """.format(itm.purchase_order),as_dict=1,debug=0)

			for row in fg_purchase_orders:
				if row.get("fg_item") in ohs:
					if row.get("pending_qty") <= ohs.get(row.get("fg_item")):
						qty_will_be_delivered = row.get("pending_qty")
						remaining_qty = flt(ohs.get(row.get("fg_item"))) - flt(row.get("pending_qty"))
						ohs.update({row.get("fg_item"):remaining_qty})
						row['qty_can_be_pulled'] = qty_will_be_delivered
					elif row.get("pending_qty") > ohs.get(row.get("fg_item")) and ohs.get(row.get("fg_item")) >= 0: 
						qty_will_be_delivered = ohs.get(row.get("fg_item"))
						ohs.update({row.get("fg_item"):0})
						row['qty_can_be_pulled'] = qty_will_be_delivered
				else : 
					qty_will_be_delivered = 0
					row['qty_can_be_pulled'] = qty_will_be_delivered
				
				self.append('pick_list_purchase_order_table',{
					'purchase_order':row.get('name'),
					'item':row.get('fg_item'),
					'qty':row.get('pending_qty'),
					'qty_can_be_pulled' :row.get('qty_can_be_pulled')
				})

		return True

	#Pick Purchase Order Pick Qty From Batch Basis On FIFO
	@frappe.whitelist()
	def get_purchase_order_items(self):
		final_raw_item_list = []
		final_data = []
		if self.pick_list_purchase_order_table:
			for item in self.pick_list_purchase_order_table:
				# Get all the avaialble locations for required items
				if self.purpose == 'Material Transfer for Subcontracted Goods':
					# final_raw_item_list.append({'item_code':item.item, 'required_qty':item.qty, 'purchase_order':item.purchase_order})
					po_raw_data = frappe.db.sql("""SELECT si.main_item_code,si.rm_item_code as item_code, si.required_qty, si.parent as sub_order,si.name as sub_order_item ,so.purchase_order from `tabSubcontracting Order Supplied Item` si join `tabSubcontracting Order` so on so.name = si.parent where so.purchase_order='{0}' and si.main_item_code='{1}' """.format(item.purchase_order, item.item), as_dict=1,debug=1)
					final_raw_item_list.extend(po_raw_data)
		
		for row in final_raw_item_list:
			item_locations_dict = get_so_and_po_item_locations(self, row.get('item_code'), self.company, row.get('required_qty'),row.get('purchase_order'), row.get("main_item_code"),row.get('sub_order_item'))
			final_data.append(item_locations_dict)
		
		for item in final_data:
			for row in item:
				item_data = get_item_defaults(row, self.company)
				for itm in item.get(row):
					self.append("purchase_order_pick_list_item",{
						"main_item": itm.get("main_item_code"),
						"item_code": row,
						"item_name": item_data.get("item_name"),
						"warehouse":itm.get("warehouse"),
						"required_qty":itm.get("required_qty"),
						"stock_qty": itm.get("qty"),
						"purchase_order" : itm.get('order'),
						"picked_qty" : itm.get('picked_qty'),
						"uom" : item_data.get('stock_uom'),
						"stock_uom":item_data.get('stock_uom'),
						"description" :item_data.get('description'),
						"item_group" : item_data.get("item_group")
					})
		
		return self.po_batch_assignment_fifo()

	def po_batch_assignment_fifo(self):
		item_list = [item.item_code for item in self.purchase_order_pick_list_item]
		wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
		item_list =  '(' + ','.join("'{}'".format(i) for i in item_list) + ')' if item_list else ()

		# batch_data = frappe.db.sql("""SELECT b.name,b.item, `tabStock Ledger Entry`.warehouse, sum(`tabStock Ledger Entry`.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` ignore index (item_code, warehouse) on (b.name = `tabStock Ledger Entry`.batch_no ) where `tabStock Ledger Entry`.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) and `tabStock Ledger Entry`.warehouse != '{1}'  group by batch_id order by b.expiry_date ASC, b.creation ASC""".format(item_list, wip_warehouse),as_dict=1,debug=1)

		batch_data = frappe.db.sql("""SELECT b.name,b.item,sle.warehouse, sum(sle.actual_qty) as qty from `tabBatch` b ,`tabSerial and Batch Entry` sbe, `tabStock Ledger Entry` sle where sle.serial_and_batch_bundle = sbe.parent and sbe.batch_no = b.name and sle.item_code in {0}  and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) and sle.warehouse != '{1}'  group by b.name order by b.expiry_date ASC, b.creation ASC""".format(item_list, wip_warehouse),as_dict=1,debug=1)

		# print("=dddddd", batch_data)
		for row in batch_data:
			calculate_remaining_po_batch_qty(row)
		
		po_list = []
		for row in self.purchase_order_pick_list_item:
			po_list.append({"main_item":row.main_item, "item_code":row.item_code, "uom":row.uom, "uom_conversion_factor":row.uom_conversion_factor, "stock_uom":row.stock_uom, "serial_nos":row.serial_no, "warehouse":row.warehouse, "required_qty":row.required_qty, "purchase_order":row.purchase_order, "stock_qty":0, "picked_qty":1})
		
		new_list=[]
		allocated_item_dict = dict()
		allocated_war_item_dict = dict()
		for row in po_list:
			po_item = row.get("item_code")+row.get("purchase_order")
			for col in batch_data:
				if row.get("item_code")==col.get("item") and row.get("warehouse")==col.get("warehouse") and po_item not in allocated_item_dict:
					item_qty=sum([r.get("picked_qty") for r in new_list if new_list and row.get("item_code")==r.get("item_code") and r.get("item_code")+r.get("purchase_order")==po_item])
					rem_reqd_qty = abs(item_qty-row.get("required_qty"))
					row.update({"required_qty":rem_reqd_qty})

					serial_no = ""
					if frappe.get_cached_value('Item', row.get("item_code"), 'has_serial_no') == 1:
						serial_nos = get_serial_no_batchwise(row.get("item_code"),col.get("name"), row.get("warehouse"), row.get("required_qty"))
						serial_no = serial_nos
					
					if col.get("qty")>= row.get("required_qty"):
						batch_qty = col.get("qty") - row.get("required_qty")
						
						new_list.append({
							"main_item":row.get("main_item"),
							"item_code":row.get("item_code"), 
							"warehouse":row.get("warehouse"), 
							"required_qty":row.get("required_qty"),
							"purchase_order":row.get("purchase_order"),
							"stock_qty":col.get("qty"),
							"picked_qty":row.get("required_qty"),
							"batch_no": col.get("name"),
							"uom":row.get("uom"), 
							"uom_conversion_factor":row.get("uom_conversion_factor"),
							"stock_uom":row.get("stock_uom"),
							"serial_no": serial_no
						})
						col.update({"qty":batch_qty})
						allocated_item_dict.update({po_item:po_item})
					else:
						reqd_qty = row.get("required_qty")-col.get("qty")
						
						if col.get("qty") > 0:
							
							new_list.append({
								"main_item":row.get("main_item"),
								"item_code":row.get("item_code"), 
								"warehouse":row.get("warehouse"), 
								"required_qty":row.get("required_qty"),
								"purchase_order":row.get("purchase_order"),
								"stock_qty":col.get("qty"),
								"picked_qty":col.get("qty"),
								"batch_no": col.get("name"),
								"uom":row.get("uom"), 
								"uom_conversion_factor":row.get("uom_conversion_factor"),
								"stock_uom":row.get("stock_uom"),
								"serial_no": serial_no
							})
						col.update({"qty":0})
		
		return new_list
# old code
	# @frappe.whitelist()
	# def new_create_dn_for_so(self):
		# item_wise_so_data =frappe.db.sql("""SELECT sales_order,item,qty from `tabPick List Sales Order Table` where parent = '{0}'""".format(self.name),as_dict=1)
		# print("============ite",item_wise_so_data)
	# 	item_code_list = [i.item for i in item_wise_so_data]
	# 	if item_wise_so_data:
	# 		final_so_dict = dict()
	# 		for row in item_wise_so_data:
	# 			if row.get("sales_order") in final_so_dict:
	# 				updated_data = final_so_dict.get(row.get("sales_order"))
	# 				updated_data.append({'sales_order':row.get('sales_order'),'item':row.get('item')})
	# 			else:
	# 				final_so_dict[row.get("sales_order")] = [{'sales_order':row.get('sales_order'),'item':row.get('item')}]
	# 	print("---------------",final_so_dict)
	# 	if final_so_dict:
	# 		for so in final_so_dict:
	# 			doc = frappe.new_doc("Delivery Note")
	# 			so_doc = frappe.get_doc("Sales Order", so)
	# 			print("============so_doc",so_doc)
	# 			doc.naming_series = "MAT-DN-.YYYY.-"
	# 			doc.customer = so_doc.customer
	# 			doc.woocommerce_order_id = so_doc.woocommerce_order_id
	# 			doc.woocommerce_order_id = so_doc.woocommerce_order_id
	# 			doc.currency = so_doc.currency
	# 			doc.selling_price_list = so_doc.selling_price_list
	# 			doc.selling_price_list = so_doc.selling_price_list
	# 			doc.taxes_and_charges = "US ST 6% - RI"
	# 			doc.tc_name = so_doc.tc_name
	# 			doc.terms = so_doc.terms
	# 			if so_doc.taxes:
	# 				for row in so_doc.taxes:
	# 					doc.append("taxes", {
	# 						"charge_type": row.charge_type,
	# 						"account_head": row.account_head,	
	# 						"description": row.account_head,
	# 						"rate": row.rate
	# 					})
	# 			for item in final_so_dict.get(so):
	# 				dn_entry = frappe.db.sql("""SELECT name From `tabDelivery Note Item` where against_sales_order='{0}' and item_code='{1}'""".format(so, item.get('item')), as_dict=1)
	# 				print("=============dn_entry",dn_entry)
	# 				if len(dn_entry) == 0:
	# 					so_data = frappe.db.sql("""SELECT item_code, warehouse, sales_order, picked_qty, batch_no, serial_no,sales_order_item from `tabSales Order Pick List Item` where sales_order='{0}' and item_code='{1}' and parent='{2}' """.format(so, item.get('item'), self.name), as_dict=1)
						
	# 					# item_code_list = "', '".join(item_code_list)
	# 					if so_data:
	# 						for row in so_doc.items:
	# 							if row.item_code == item.get('item'):
	# 								for itm in so_data:
	# 									doc.append("items", {
	# 										"item_code": row.item_code,
	# 										"item_name": row.item_name,
	# 										"description": row.description,
	# 										"uom": row.uom,
	# 										"stock_uom": row.stock_uom,
	# 										"conversion_factor": row.conversion_factor,
	# 										"qty": itm.get("picked_qty"),
	# 										"rate": row.rate,
	# 										"warehouse": itm.get("warehouse"),
	# 										"batch_no": itm.get("batch_no"),
	# 										"against_sales_order": so,
	# 										"serial_no": itm.get("serial_no"),
	# 										"so_detail":itm.get("sales_order_item")
	# 									})	
	# 			doc.save()
	# 			if doc.name:
	# 				status = "To Bill" if doc.docstatus==1 else "Draft"
	# 				if item_code_list:
	# 					print("=============item",item_code_list)
	# 					for item in item_code_list:
	# 						frappe.db.set_value("Pick List Sales Order Table",{'parent':self.name,'item':item,'sales_order':so},'delivery_note',doc.name)
	# 						frappe.db.set_value("Pick List Sales Order Table",{'parent':self.name,'item':item,'sales_order':so},'delivery_note_status',doc.status)
	# 				frappe.msgprint("Delivery Note created {0}".format(doc.name))
					
	# 			else:
	# 				frappe.msgprint("Delivery Note already created for Sales Order {0}".format(so.sales_order))
# old code end 
	 

# new code 
	@frappe.whitelist()
	def new_create_dn_for_so(self):
		item_wise_so_data = frappe.db.sql("""
			SELECT sales_order, item, qty 
			FROM `tabPick List Sales Order Table` 
			WHERE parent = '{0}'
		""".format(self.name), as_dict=1)


		item_code_list = [i.item for i in item_wise_so_data]

		if item_wise_so_data:
			final_so_dict = dict()

			for row in item_wise_so_data:
				sales_order = row.get("sales_order")
				if sales_order in final_so_dict:
					final_so_dict[sales_order].append({'sales_order': sales_order, 'item': row.get('item')})
				else:
					final_so_dict[sales_order] = [{'sales_order': sales_order, 'item': row.get('item')}]


		if final_so_dict:
			for so in final_so_dict:
				doc = frappe.new_doc("Delivery Note")
				so_doc = frappe.get_doc("Sales Order", so)

				 
				doc.naming_series = "MAT-DN-.YYYY.-"
				doc.customer = so_doc.customer
				doc.woocommerce_order_id = so_doc.woocommerce_order_id
				doc.currency = so_doc.currency
				doc.selling_price_list = so_doc.selling_price_list
				doc.taxes_and_charges = "US ST 6% - RI"
				doc.tc_name = so_doc.tc_name
				doc.terms = so_doc.terms

				if so_doc.taxes:
					for row in so_doc.taxes:
						doc.append("taxes", {
							"charge_type": row.charge_type,
							"account_head": row.account_head,
							"description": row.account_head,
							"rate": row.rate or 0  
						})

				for item in final_so_dict.get(so):
					dn_entry = frappe.db.sql("""
						SELECT name 
						FROM `tabDelivery Note Item` 
						WHERE against_sales_order='{0}' AND item_code='{1}'
					""".format(so, item.get('item')), as_dict=1)


					if not dn_entry: 
						so_data = frappe.db.sql("""
							SELECT item_code, warehouse, sales_order, picked_qty, batch_no, serial_no, sales_order_item 
							FROM `tabSales Order Pick List Item` 
							WHERE sales_order='{0}' AND item_code='{1}' AND parent='{2}'
						""".format(so, item.get('item'), self.name), as_dict=1)

						if so_data:
							for row in so_doc.items:
								if row.item_code == item.get('item'):
									for itm in so_data:
										picked_qty = itm.get("picked_qty") or 1
										doc.append("items", {
											"item_code": row.item_code,
											"item_name": row.item_name,
											"description": row.description,
											"uom": row.uom,
											"stock_uom": row.stock_uom,
											"conversion_factor": row.conversion_factor,
											"qty": picked_qty,
											# "rate": row.rate or 0, 
											# "base_rate" : row.rate or 0,
											# "base_amount":picked_qty * (row.rate if row.rate is not None else 0),
											"warehouse": itm.get("warehouse"),
											"batch_no": itm.get("batch_no"),
											"against_sales_order": so,
											"serial_no": itm.get("serial_no"),
											"so_detail": itm.get("sales_order_item")
										})

				# Save the document
				doc.save()
				if doc.name:
					# frappe.throw(str(doc.name))
					frappe.msgprint("Delivery Note created: {0}".format(doc.name))
					
					if item_code_list:
						for item in item_code_list:
							frappe.db.set_value("Pick List Sales Order Table", {'parent': self.name, 'item': item, 'sales_order': so}, 'delivery_note', doc.name)
							frappe.db.set_value("Pick List Sales Order Table", {'parent': self.name, 'item': item, 'sales_order': so}, 'delivery_note_status', doc.status)
				else:
					frappe.msgprint("Delivery Note already created for Sales Order {0}".format(so))


# new code end



		# for so in self.pick_list_sales_order_table:
		# 	dn_entry = frappe.db.sql("""SELECT name From `tabDelivery Note Item` where against_sales_order='{0}' and item_code='{1}'""".format(so.sales_order, so.item), as_dict=1)
			
		# 	if so.select_item_for_delivery and len(dn_entry)==0:
		# 		so_data = frappe.db.sql("""SELECT item_code, warehouse, sales_order, picked_qty, batch_no, serial_no,sales_order_item from `tabSales Order Pick List Item` where sales_order='{0}' and item_code='{1}' and parent='{2}' """.format(so.sales_order, so.item, so.parent), as_dict=1)
		# 		if so_data:

		# 			doc = frappe.new_doc("Delivery Note")

		# 			so_doc = frappe.get_doc("Sales Order", so.sales_order)
		# 			doc.naming_series = "MAT-DN-.YYYY.-"
		# 			doc.customer = so_doc.customer
		# 			doc.woocommerce_order_id = so_doc.woocommerce_order_id
		# 			doc.woocommerce_order_id = so_doc.woocommerce_order_id
		# 			doc.currency = so_doc.currency
		# 			doc.selling_price_list = so_doc.selling_price_list
		# 			doc.selling_price_list = so_doc.selling_price_list
		# 			doc.taxes_and_charges = "US ST 6% - RI"
		# 			doc.tc_name = so_doc.tc_name
		# 			doc.terms = so_doc.terms
		# 			for row in so_doc.items:
		# 				if row.item_code == so.item:
		# 					for itm in so_data:
		# 						doc.append("items", {
		# 							"item_code": row.item_code,
		# 							"item_name": row.item_name,
		# 							"description": row.description,
		# 							"uom": row.uom,
		# 							"stock_uom": row.stock_uom,
		# 							"conversion_factor": row.conversion_factor,
		# 							"qty": itm.get("picked_qty"),
		# 							"rate": row.rate,
		# 							"warehouse": itm.get("warehouse"),
		# 							"batch_no": itm.get("batch_no"),
		# 							"against_sales_order": so.sales_order,
		# 							"serial_no": itm.get("serial_no"),
		# 							"so_detail":itm.get("sales_order_item")
		# 						})

		# 			if not so_doc.taxes:
		# 				for row in so_doc.taxes:
		# 					doc.append("taxes", {
		# 						"charge_type": row.charge_type,
		# 						"account_head": row.account_head,	
		# 						"description": row.account_head,
		# 						"rate": row.rate
		# 					})
		# 			doc.save()
		# 			if doc.name:
		# 				status = "To Bill" if doc.docstatus==1 else "Draft"
		# 				frappe.msgprint("Delivery Note created {0}".format(doc.name))
		# 				for row in doc.items: 
		# 					frappe.db.set_value("Pick List Sales Order Table", {"sales_order":row.against_sales_order,"item":row.item_code}, {"delivery_note":doc.name, "delivery_note_status":status})
		# 					frappe.db.commit()
		# 					return {"name":doc.name}
		# 		else:
		# 			frappe.msgprint("Please Pick Quantity For Atleast One Item for Sales Order {0}".format(so.sales_order))
		# 	elif not so.select_item_for_delivery:
		# 		frappe.msgprint("Please Select Item For Delivery For Row {0}".format(so.idx))
		# 	else:
		# 		frappe.msgprint("Delivery Note already created for Sales Order {0}".format(so.sales_order))
	
# def get_available_item_locations_for_batched_item(item_code, from_warehouses, required_qty, company):
def get_available_item_locations_for_batched_item(item_list):
	joined_item_list = "', '".join(item_list)

	# warehouse_condition = "and warehouse in %(warehouses)s" if from_warehouses else ""
	batch_locations = frappe.db.sql(
		"""
		SELECT
			sle.`warehouse`,
			sle.`batch_no`,
			SUM(sle.`actual_qty`) AS `qty`
		FROM
			`tabStock Ledger Entry` sle, `tabBatch` batch
		WHERE
			sle.batch_no = batch.name
			and sle.`item_code` in %(('{item_code}'))s
			and batch.disabled = 0
			and sle.is_cancelled=0
			and IFNULL(batch.`expiry_date`, '2200-01-01') > %(today)s
		GROUP BY
			sle.`batch_no`
		HAVING `qty` > 0
		ORDER BY IFNULL(batch.`expiry_date`, '2200-01-01'), batch.`creation`
	""",{
			"today": today(),
			"item_code":"', '".join(item_list)
		},
		as_dict=1,debug=1
	)

	return batch_locations

def calculate_remaining_batch_qty(row):
	used_qty = frappe.db.sql("""SELECT item_code, sum(picked_qty) as picked_qty From `tabConsolidated Pick List` co join `tabWork Order Pick List Item` woi on woi.parent = co.name where batch_no='{0}' and item_code='{1}' and co.purpose = 'Manufacture By FG' and co.docstatus=1""".format(row.name, row.item), as_dict=1,debug=0)
	if used_qty:
		row["qty"] = (row.qty - flt(used_qty[0].get('picked_qty'))) if (row.qty - flt(used_qty[0].get('picked_qty'))) > 0 else 0

def calculate_remaining_so_batch_qty(row):
	used_qty = frappe.db.sql("""SELECT soi.item_code, sum(soi.picked_qty) as picked_qty From `tabSales Order Pick List Item` soi join `tabConsolidated Pick List` co on co.name = soi.parent where soi.batch_no='{0}' and soi.item_code='{1}' and co.docstatus=0""".format(row.name, row.item), as_dict=1)
	if used_qty:
		row["qty"] = abs(row.qty - flt(used_qty[0].get('picked_qty')))

def calculate_remaining_po_batch_qty(row):
	used_qty = frappe.db.sql("""SELECT item_code, sum(picked_qty) as picked_qty From `tabPurchase Order Pick List Item` where batch_no='{0}' and item_code='{1}' and docstatus=1""".format(row.name, row.item), as_dict=1)
	if used_qty:
		row["qty"] = abs(row.qty - flt(used_qty[0].get('picked_qty')))

def get_serial_nos(item_code,batch_id):
	serial_nos = frappe.db.sql("""SELECT name from `tabSerial No` where batch_no = '{0}' and item_code = '{1}'""".format(batch_id,item_code),as_dict=1)
	serial_nos = [item.name for item in serial_nos]
	return serial_nos
def get_serial_no_batchwise(item_code,batch_no,warehouse,qty):
	return ", ".join(frappe.db.sql_list("""SELECT name from `tabSerial No`
		where item_code=%(item_code)s and warehouse=%(warehouse)s 
		and batch_no=IF(%(batch_no)s IS NULL, batch_no, %(batch_no)s) order
		by timestamp(purchase_date, purchase_time) asc limit %(qty)s""", {
			"item_code": item_code,
			"warehouse": warehouse,
			"batch_no": batch_no,
			"qty": abs(cint(qty))
		}, debug=0))

def get_item_locations(self,item_list,company):
	item_locations_dict = dict()
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", "default_wip_warehouse")
	warehouses = [x.get('name') for x in frappe.get_list("Warehouse", {'company': company}, "name")]
	if wip_warehouse in warehouses:
		warehouses.remove(wip_warehouse)
	for item in item_list :
		item_locations = frappe.get_all('Bin',
							fields=['warehouse', 'actual_qty as qty'],
							filters={'item_code':item,
							'actual_qty': ['>', 0],
							'warehouse' :['in', warehouses]
							},
							order_by='creation')
		item_locations_dict[item] = item_locations
	return item_locations_dict

def get_so_and_po_item_locations(self, item, company, qty, order, main_item_code,sales_order_item):
	item_locations_dict = dict()
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", "default_wip_warehouse")
	warehouses = [x.get('name') for x in frappe.get_list("Warehouse", {'company': company}, "name")]
	# if wip_warehouse in warehouses:
	# 	warehouses.remove(wip_warehouse)
	if item:
		item_locations = frappe.get_all('Bin',
							fields=['warehouse', 'actual_qty as qty'],
							filters={'item_code':item,
							'actual_qty': ['>', 0],
							'warehouse' :['in', warehouses]
							},
							order_by='creation')

		for row in item_locations:
			row["required_qty"] = qty
			row["order"] = order
			row["main_item_code"] = main_item_code
			row['sales_order_item'] = sales_order_item
		item_locations_dict[item] = item_locations
	return item_locations_dict

def get_item_locations_for_manufacture(self,item_list,company):
	item_locations_dict = dict()
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", "default_wip_warehouse")
	# warehouses = [x.get('name') for x in frappe.get_list("Warehouse", {'company': company}, "name")]
	# if wip_warehouse in warehouses:
	# 	warehouses.remove(wip_warehouse)
	for item in item_list :
		item_locations = frappe.get_all('Bin',
							fields=['warehouse', 'actual_qty as qty'],
							filters={'item_code':item,
							'actual_qty': ['>', 0],
							'warehouse' :['=', wip_warehouse]
							},
							order_by='creation')
		item_locations_dict[item] = item_locations
	return item_locations_dict	
def get_raw_material(bom_no, company, qty,work_order):
	try:
		work_order_doc = frappe.get_doc("Work Order",work_order)
		if work_order_doc.use_multi_level_bom == 1:
			raw_materials = get_bom_items_as_dict(bom_no,company,qty,fetch_exploded=1)
		else:
			raw_materials = get_bom_items_as_dict(bom_no,company,qty,fetch_exploded=0)
		for item in raw_materials:
			item_info = raw_materials.get(item)
			item_info['work_order'] = work_order
		return raw_materials
	except Exception as e:
		frappe.log_error(message = frappe.get_traceback() , title = "Get Raw Material")

@frappe.whitelist()
def get_work_orders(production_plan):
	if production_plan:
		work_order_data = frappe.db.sql("""SELECT wo.name,wo.qty,wo.produced_qty,(wo.qty-wo.produced_qty) as pending_qty from `tabWork Order` wo where production_plan = '{0}' and (wo.produced_qty < wo.qty)  and wo.status in ('In process','Not Started') and wo.docstatus = 1 order by wo.bom_level asc""".format(production_plan),as_dict =1,debug=0)
		return work_order_data

@frappe.whitelist()
def get_work_order_data(work_order):
	if work_order:
		work_order_data = frappe.db.sql("""SELECT wo.name,wo.qty,wo.produced_qty,(wo.qty-wo.produced_qty) as pending_qty from `tabWork Order` wo where wo.name = '{0}' and (wo.produced_qty < wo.qty)  and wo.status in ('In process','Not Started') and wo.docstatus = 1 order by wo.name asc""".format(work_order),as_dict =1,debug=0)
		doc = frappe.get_doc("Work Order",work_order)
		ohs = get_current_stock()
		qty_will_be_produced_list = []
		for item in doc.required_items:
			if item.item_code in ohs:
				if item.required_qty <= ohs.get(item.item_code):
					percent_stock = 100
					qty_will_be_produced = item.required_qty
					qty_will_be_produced_list.append(qty_will_be_produced)
				elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = (percent_stock/100*item.required_qty)
					qty_will_be_produced_list.append(qty_will_be_produced)
				else : 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = 0
					qty_will_be_produced_list.append(qty_will_be_produced)
		
		work_order_data[0]['qty_will_be_produced'] =min(qty_will_be_produced_list)
		return work_order_data
def get_wip_stock():
	# 1.get wip warehouse
	# item_list = "', '".join(item_list) 
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", 'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse = '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict
def get_raw_bom_data(bom):
	if bom:
		raw_bom_data = frappe.db.sql("""SELECT b.item as item_code,b.quantity,boi.item_code,boi.qty as qty from `tabBOM` b join `tabBOM Item` boi on b.name = boi.parent  join `tabItem` i on i.item_code = boi.item_code where b.name = '{0}' and i.is_stock_item = 1 """.format(bom),as_dict=1)
		if raw_bom_data:
			return raw_bom_data
def get_current_stock():
	# 1.get wip warehouse
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", 'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse != '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict
@frappe.whitelist()
def get_child_boms(bom,bom_childs):
	if bom:
		childs = frappe.db.sql("""SELECT bom_no as bom from `tabBOM Item` where parent = '{0}' and bom_no != '' """.format(bom),as_dict=1,debug=1)
		# print("================childs",childs)
		bom_childs += childs

		if len(childs)>0:
			for c in childs:
				if c.get('bom'):
					get_child_boms(c.get('bom'),bom_childs)
		return bom_childs
@frappe.whitelist()
def get_all_boms_in_order(bom_childs):
	if len(bom_childs)>=1:
		final_list = [row.get('bom') for row in bom_childs if row.get("bom")]
		final_list = '(' + ','.join("'{}'".format(i) for i in final_list) + ')'
		childs = frappe.db.sql("""SELECT b.name as bom,b.bom_level from `tabBOM` b where b.name in {0} order by b.bom_level asc""".format(final_list),as_dict=1)
		return childs
def get_current_stock_for_manufacture():
	# 1.get wip warehouse
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings", 'default_wip_warehouse')
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse = '{0}' group by item_code """.format(wip_warehouse),as_dict=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict	
@frappe.whitelist()
def validate_picked_qty(work_order,required_qty,picked_qty,doc_name,row_name,item_code):
	picked_qty = frappe.db.sql("""SELECT sum(picked_qty) as picked_qty from `tabWork Order Pick List Item` where parent = '{0}' and work_order = '{1}' and item_code = '{2}' and name != '{3}'""".format(doc_name,work_order,item_code,row_name),as_dict=1)
	return picked_qty[0].get('picked_qty')

@frappe.whitelist()
def check_stock_entries(work_order,consolidated_pick_list):
	if work_order and consolidated_pick_list:
		stock_entry_list = frappe.db.sql("""SELECT name from `tabStock Entry` where work_order = '{0}' and consolidated_pick_list = '{1}' and docstatus != '2'""".format(work_order,consolidated_pick_list),as_dict=1)
		if len(stock_entry_list)==0:
			# check dependent work orders
			wo_doc = frappe.get_doc("Work Order",work_order)
			if wo_doc:
				required_items = frappe.db.sql("SELECT i.item_code ,b.name from `tabWork Order Item` i join `tabBOM` b on b.item = i.item_code where i.parent = '{0}' and b.is_default = 1 and b.is_Active =1".format(work_order),as_dict=1)
				
				# required_items = "', '".join(required_items)

			return True
		else:
			frappe.throw("Stock Entries already created for work order {0}".format(work_order))
@frappe.whitelist()
def create_stock_entry(work_order, consolidated_pick_list, row_name):
	try:
		if work_order and consolidated_pick_list:
			work_order_doc = frappe.get_doc("Work Order",work_order)
			pick_list_doc = frappe.get_doc("Consolidated Pick List",consolidated_pick_list)
			work_order_list = [wo.work_order for wo in pick_list_doc.work_orders]
			work_order_list = "', '".join(work_order_list)
			qty_of_finish_good = frappe.db.get_value("Pick Orders",{'parent':consolidated_pick_list,'work_order':work_order},'qty_of_finished_goods_to_pull')
			if pick_list_doc.purpose in ['Manufacture By FG','Manufacture By Work Order']:
				# check_dependent_wo = frappe.db.sql("""SELECT wo.name from `tabWork Order` wo where wo.production_planning_with_lead_time = '{0}' and wo.so_reference = '{1}' and wo.produced_qty = 0 and wo.bom_level < {2} and wo.parent_item_code = '{3}'""".format(work_order_doc.production_planning_with_lead_time,work_order_doc.so_reference,work_order_doc.bom_level,work_order_doc.parent_item_code),as_dict=1)
				check_dependent_wo = frappe.db.sql("""SELECT wo.name from `tabWork Order` wo where wo.production_planning_with_lead_time = '{0}' and wo.so_reference = '{1}' and wo.produced_qty = 0 and wo.bom_level < {2} and wo.parent_item_code = '{3}' and wo.status != 'Completed' and wo.name in ('{4}')""".format(work_order_doc.production_planning_with_lead_time,work_order_doc.so_reference,work_order_doc.bom_level,work_order_doc.parent_item_code,work_order_list),as_dict=1,debug=1)
				check_dependent_wo = [i.name for i in check_dependent_wo]
				if check_dependent_wo and len(check_dependent_wo):
					frappe.throw("Please Complete Dependent Work Order Before Proceeding".format(check_dependent_wo))
				job_cards = frappe.db.sql("""SELECT name from `tabJob Card` where work_order = '{0}' and docstatus = 0""".format(work_order),as_dict=1)
				if job_cards:
					for row in job_cards:
						job_card = frappe.get_doc("Job Card",row.get('name'))
						if job_card:
							job_card.total_completed_qty = job_card.for_quantity
							job_card.append('time_logs',{
								'from_time':nowdate(),
								'to_time':nowdate(),
								'completed_qty':qty_of_finish_good})
							job_card.save()
							job_card.submit()
			data =  frappe.db.sql("""SELECT item_code,warehouse as s_warehouse,picked_qty,work_order,stock_uom,engineering_revision,batch_no,serial_no from `tabWork Order Pick List Item` where parent = '{0}' and work_order = '{1}' and picked_qty > 0""".format(consolidated_pick_list,work_order),as_dict=1, debug=0)
			items = [i.item_code for i in data]
			work_order_items = frappe.db.sql("""SELECT wo.item_code from `tabWork Order Item` wo where wo.parent = '{0}'""".format(work_order),as_dict=1,debug=1)
			if work_order_items:
				work_order_items = [i.item_code for i in work_order_items]
			remaining_items = []
			wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
			for item in work_order_items:
				if item not in items:
					remaining_items.append(item)
			# wip_data = frappe.db.sql("""SELECT """)
			# if len(data) > 0:
			stock_entry = frappe.new_doc("Stock Entry")
			if stock_entry:
				if pick_list_doc.purpose in ['Material Transfer for Manufacture By FG','Material Transfer for Manufacture By Work Order']:
					stock_entry.stock_entry_type = 'Material Transfer for Manufacture'
					stock_entry.company = work_order_doc.company
					stock_entry.work_order = work_order
					stock_entry.consolidated_pick_list = consolidated_pick_list
					stock_entry.from_bom = 1
					stock_entry.bom_no = work_order_doc.bom_no
					stock_entry.fg_completed_qty = qty_of_finish_good
					# stock_entry.fg_completed_qty = data.get("fg_completed_qty")
					# stock_entry.from_warehouse = data.get("selected_warehouse")
					stock_entry.to_warehouse = work_order_doc.wip_warehouse
					if len(data) > 0:
						for row in data:
							stock_entry.append("items",{
									"item_code": row.get("item_code"),
										"qty": row.get("picked_qty"),
										"s_warehouse": row.get("s_warehouse"),
										"t_warehouse": work_order_doc.wip_warehouse,
										'batch_no' : row.get("batch_no"),
										'serial_no' : row.get("serial_no")
								})
						stock_entry.save()
					else:
						frappe.throw("Stock Available for work order {0} in warehouse {1}.Remove {0} from Pick List Work Order Table".format( work_order_doc.name,work_order_doc.wip_warehouse))
				else:
					stock_entry.stock_entry_type = 'Manufacture'
					stock_entry.company = work_order_doc.company
					stock_entry.work_order = work_order
					stock_entry.consolidated_pick_list = consolidated_pick_list
					stock_entry.from_bom = 1
					stock_entry.bom_no = work_order_doc.bom_no
					stock_entry.fg_completed_qty = qty_of_finish_good
					# stock_entry.fg_completed_qty = data.get("fg_completed_qty")
					# stock_entry.from_warehouse = data.get("selected_warehouse")
					stock_entry.to_warehouse = work_order_doc.fg_warehouse
					if len(data) > 0:
						for row in data:
							stock_entry.append("items",{
								"item_code": row.get("item_code"),
									"qty": row.get("picked_qty"),
									"s_warehouse": row.get("s_warehouse"),
									# "t_warehouse": work_order_doc.fg_warehouse,
									'batch_no' : row.get("batch_no"),
									'serial_no' : row.get("serial_no")

							})
					if remaining_items:
						joined_item_list = "', '".join(remaining_items)
						# batch_data = frappe.db.sql("""SELECT b.name,b.item, sle.warehouse, sum(sle.actual_qty) as qty from `tabBatch` b join `tabStock Ledger Entry` sle on (b.name = sle.batch_no ) where sle.item_code in ('{0}')  and IFNULL(b.expiry_date, '2200-01-01') > %(today)s and sle.warehouse = '{1}' and b.disabled = 0 and sle.is_cancelled=0 group by b.name,sle.warehouse HAVING qty > 0 order by  b.creation ASC""".format(joined_item_list,wip_warehouse),{'today':today()},as_dict=1,debug=1)
						batch_data  = frappe.db.sql(
							"""
							SELECT
								sle.`warehouse` as warehouse,
								sle.`batch_no` as name,
								SUM(sle.`actual_qty`) AS `qty`,
								sle.item_code as item
							FROM
								`tabStock Ledger Entry` sle, `tabBatch` batch
							WHERE
								sle.batch_no = batch.name
								and sle.`item_code` in ('{0}')
								and batch.disabled = 0
								and sle.is_cancelled=0
								and sle.warehouse = '{1}'
							GROUP BY
								sle.`batch_no`
							HAVING `qty` > 0
							ORDER BY IFNULL(batch.`expiry_date`, '2200-01-01'), batch.`creation`
						""".format(joined_item_list,wip_warehouse),
							as_dict=1,debug=1
						)
						if batch_data:
							for row in batch_data:
							# 	calculate_remaining_batch_qty(row)
								
								if row.item in remaining_items:
									wo_qty = frappe.db.get_value("Work Order Item",{'parent':work_order,'item_code':row.item},'required_qty')
									if row.qty >= wo_qty:
										stock_entry.append("items",{
											"item_code": row.item,
											"qty": wo_qty,
											"s_warehouse": row.warehouse,
											# "t_warehouse": work_order_doc.fg_warehouse,
											'batch_no' : row.name,
										})
										pick_list_doc.append('work_order_pick_list_item',{
											'item_code':row.item,
											# 'item_name':row.item_name,
											# 'item_group':row.item_group,
											'work_order':work_order,
											# 'uom':row.uom,
											# 'stock_uom':row.stock_uom,
											'batch_no':row.name,
											'picked_qty':row.qty,
											'warehouse':row.warehouse
										})
									else:
										stock_entry.append("items",{
											"item_code": row.item,
											"qty": row.qty,
											"s_warehouse": row.warehouse,
											# "t_warehouse": work_order_doc.fg_warehouse,
											'batch_no' : row.name,
										})
										pick_list_doc.append('work_order_pick_list_item',{
											'item_code':row.item,
											# 'item_name':row.item_name,
											# 'item_group':row.item_group,
											'work_order':work_order,
											# 'uom':row.uom,
											# 'stock_uom':row.stock_uom,
											'batch_no':row.name,
											'picked_qty':row.qty,
											'warehouse':row.warehouse
										})
					stock_entry.append("items",{
						"item_code":work_order_doc.production_item,
						"qty": qty_of_finish_good,
						"t_warehouse": work_order_doc.fg_warehouse
					})	
					stock_entry.save()
					pick_list_doc.save()
					if stock_entry.name:
						sub_assembly_data = frappe.db.sql("""SELECT item_code,item_name,item_group,stock_uom,uom,t_warehouse,work_order,batch_no from `tabStock Entry` se join `tabStock Entry Detail` sed on sed.parent = se.name where se.name = '{0}' and t_warehouse is not null""".format(stock_entry.name),as_dict=1,debug=1)
				frappe.db.set_value("Pick Orders", {"name":row_name}, {"stock_entry":stock_entry.name, "stock_entry_status":"Draft"})
				frappe.db.commit()
					
				return stock_entry.name
			# else:
			# 	frappe.throw("Please Pick Quantity For Atleast One Item")
	except Exception as e:
		traceback = frappe.get_traceback()
		frappe.log_error(
			title=_("Error while creating stock Entry"),
			message=traceback,
		)
		raise e
def wo_filter_condition(filters):
	conditions = "1=1"
	if filters.get("work_order"):
		conditions += " and wo.name = '{0}'".format(filters.get("work_order"))
	# if filters.get("customer"):
	# 	conditions+= " and wo. = '{0}'".format(filters.get("customer"))
	if filters.get("planned_start_date"):
		conditions += " and wo.planned_start_date >= '{0}'".format(filters.get('planned_start_date'))
	if filters.get("planned_end_date"):
		conditions += " and wo.planned_end_date <= '{0}'".format(filters.get('planned_end_date'))
	if filters.get("expected_delivery_date"):
		conditions += " and wo.expected_delivery_date = '{0}'".format(filters.get('expected_delivery_date'))
	if filters.get("production_planning_with_lead_time"):
		conditions += " and wo.production_planning_with_lead_time = '{0}'".format(filters.get('production_planning_with_lead_time'))
	return conditions

def so_filter_condition(filters):
	conditions = "1=1"
	if filters.get("sales_order"):
		conditions += " and so.name = '{0}'".format(filters.get("sales_order"))
	if filters.get("customer"):
		conditions += " and so.customer = '{0}'".format(filters.get("customer"))
	if filters.get("customer_po_number"):
		conditions += " and so.po_no = '{0}'".format(filters.get('customer_po_number'))
	if filters.get("delivery_start_date"):
		conditions += " and so.delivery_date >= '{0}'".format(filters.get('delivery_start_date'))
	if filters.get("delivery_end_date"):
		conditions += " and so.delivery_date <= '{0}'".format(filters.get('delivery_end_date'))
	return conditions


def po_filter_condition(filters):
	conditions = "1=1"
	if filters.get("supplier"):
		conditions += " and po.supplier = '{0}'".format(filters.get("supplier"))
	if filters.get("purchase_order"):
		conditions += " and po.name = '{0}'".format(filters.get("purchase_order"))
	if filters.get("start_date"):
		conditions += " and po.schedule_date >= '{0}'".format(filters.get('start_date'))
	if filters.get("end_date"):
		conditions += " and po.schedule_date <= '{0}'".format(filters.get('end_date'))
	return conditions


@frappe.whitelist()
def get_pick_list_details(doc):
	data = json.loads(doc)
	header = ['Sr.No','Work Order','Total Qty of Finished Goods on Work Order','Qty of Finished Goods Already Completed','Qty of Finished Goods To Pull']
	header_2 = ['Sr.No','Item','Warehouse','Work Order','Required Qty','Stock Qty','Picked Qty','Item Name','Description','Item Group','Batch','Engineering Revision','UOM','Stock UOM','Conversion Factor']
	
	book = Workbook()
	sheet = book.active
	
	row = 1
	col = 1

	cell = sheet.cell(row=row,column=col)
	cell.value = 'Company'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+1)
	cell.value = data.get("company")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+1,column=col)
	cell.value = 'Purpose'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+1,column=col+1)
	cell.value = data.get("purpose")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+2,column=col)
	cell.value = 'Production Plan'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+2,column=col+1)
	cell.value = data.get("production_plan")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row+4,column=col)
	cell.value = 'Pick List Work Order Table'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	for item in header:
		cell = sheet.cell(row=row+5,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

		col+=1
	row = 7
	col = 1
	# last_row = row
	for item in data.get("work_orders"):
		cell = sheet.cell(row=row,column=col)
		cell.value = item.get('idx')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+1)
		cell.value = item.get('work_order')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = item.get('total_qty_to_of_finished_goods_on_work_order')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+3)
		cell.value = item.get('qty_of_finished_goods_already_completed')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		cell = sheet.cell(row=row,column=col+4)
		cell.value = item.get('qty_of_finished_goods_to_pull')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		
		row+=1

	cell = sheet.cell(row=row+1,column=col)
	cell.value = 'Work Order Pick List Item'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	for item in header_2:
		cell = sheet.cell(row=row+2,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')

		col+=1
	row = row+3
	col = 1
	for item in data.get('work_order_pick_list_item'):
		cell = sheet.cell(row=row,column=col)
		cell.value = item.get('idx')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+1)
		cell.value = item.get('item_code')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+2)
		cell.value = item.get('warehouse')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+3)
		cell.value = item.get('work_order')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+4)
		cell.value = item.get('required_qty')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+5)
		cell.value = item.get('stock_qty')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+6)
		cell.value = item.get('picked_qty')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = sheet.cell(row=row,column=col+7)
		cell.value = item.get('item_name')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = sheet.cell(row=row,column=col+8)
		cell.value = item.get('description')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+9)
		cell.value = item.get('item_group')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+10)
		cell.value = item.get('batch_no')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")


		cell = sheet.cell(row=row,column=col+11)
		cell.value = item.get('engineering_revision')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+12)
		cell.value = item.get('uom')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+13)
		cell.value = item.get('stock_uom')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		cell = sheet.cell(row=row,column=col+14)
		cell.value = item.get('conversion_factor')
		cell.font = cell.font.copy(bold=False)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")

		row+=1



	file_path = frappe.utils.get_site_path("public")
	# now = datetime.now()
	fname = "WORK_ORDER_PICK_LIST" + nowdate() + ".xlsx"
	book.save(file_path+fname)


@frappe.whitelist()
def download_xlsx():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	# now = datetime.now()
	fname = "WORK_ORDER_PICK_LIST" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname



@frappe.whitelist()
def get_batch_no(item_code, warehouse, qty=1, throw=False):
	"""
	Get batch number using First Expiring First Out method.
	:param item_code: `item_code` of Item Document
	:param warehouse: name of Warehouse to check
	:param qty: quantity of Items
	:return: String represent batch number of batch with sufficient quantity else an empty String
	"""

	batch_no = None
	batches = get_batches(item_code, warehouse, qty, throw)
	for batch in batches:
		if cint(qty) <= cint(batch.qty):
			batch_no = batch.name
			break

	# if not batch_no:
	# 	frappe.msgprint(_('Please select a Batch for Item {0}. Unable to find a single batch that fulfills this requirement').format(frappe.bold(item_code)))
	# 	if throw:
	# 		raise UnableToSelectBatchError

	return batch_no


def get_batches(item_code, warehouse, qty=1, throw=False):
	# from erpnext.stock.doctype.serial_no.serial_no import get_serial_nos
	cond = ''
	if frappe.get_cached_value('Item', item_code, 'has_batch_no'):
		# serial_nos = get_serial_nos(serial_no)
		batch = frappe.get_all("Serial No",
			fields = ["distinct batch_no"],
			filters= {
				"item_code": item_code,
				"warehouse": warehouse
			}
		)
		# if not batch:
		# 	validate_serial_no_with_batch(serial_nos, item_code)

		if batch and len(batch) > 1:
			return []

		cond = " and b.name = %s" %(frappe.db.escape(batch[0].batch_no))

	return frappe.db.sql("""
		SELECT b.name, sum(`tabStock Ledger Entry`.actual_qty) as qty
		from `tabBatch` b
			join `tabStock Ledger Entry` ignore index (item_code, warehouse)
				on (b.name = `tabStock Ledger Entry`.batch_no )
		where `tabStock Ledger Entry`.item_code = %s and `tabStock Ledger Entry`.warehouse = %s
			and (b.expiry_date >= CURDATE() or b.expiry_date IS NULL) {0}
		group by batch_id
		order by b.expiry_date ASC, b.creation ASC
	""".format(cond), (item_code, warehouse), as_dict=True)

def validate_serial_no_with_batch(serial_nos, item_code):
	if frappe.get_cached_value("Serial No", serial_nos[0], "item_code") != item_code:
		frappe.throw(_("The serial no {0} does not belong to item {1}")
			.format(get_link_to_form("Serial No", serial_nos[0]), get_link_to_form("Item", item_code)))

	serial_no_link = ','.join(get_link_to_form("Serial No", sn) for sn in serial_nos)

	message = "Serial Nos" if len(serial_nos) > 1 else "Serial No"
	frappe.throw(_("There is no batch found against the {0}: {1}")
		.format(message, serial_no_link))

@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_work_order_for_pick_list(doctype, txt, searchfield, start, page_len, filters):
	fg_item_groups = frappe.db.sql("""SELECT item_group from `tabTable For Item Group`""",as_dict=1)
	todays_date = datetime.date.today()
	fg_item_groups = [item.get('item_group') for item in fg_item_groups]
	fg_item_groups = "', '".join(fg_item_groups)
	if filters.get("purpose") == "Material Transfer for Manufacture By FG":
		return frappe.db.sql(""" SELECT wo.name FROM `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.planned_start_date >= '{0}' and i.item_group in ('{1}')""".format(todays_date,fg_item_groups),debug=1)
	elif filters.get("purpose") == "Manufacture By FG":
		return frappe.db.sql(""" SELECT wo.name FROM `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item  where wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 or wo.material_transferred_for_manufacturing > 0 or wo.produced_qty < wo.qty and wo.planned_start_date >= '{0}' and i.item_group in ('{1}')""".format(todays_date,fg_item_groups),debug=1)
	elif filters.get("purpose") == "Material Transfer for Manufacture By Work Order":
		return frappe.db.sql(""" SELECT wo.name FROM `tabWork Order` wo join `tabItem` i on i.item_code = wo.production_item where wo.status in ('Not Started','In Process') and wo.skip_transfer = 0 and wo.planned_start_date >= '{0}'""".format(todays_date),debug=1)
	else:
		return frappe.db.sql(""" SELECT wo.name FROM `tabWork Order` wo join `tabItem` i on i.item_code == wo.production_item  where wo.status in ('Not Started','In Process') and wo.skip_transfer = 1 or wo.material_transferred_for_manufacturing > 0 or wo.produced_qty < wo.qty and wo.planned_start_date >= '{0}'""".format(todays_date),debug=1)

@frappe.whitelist()
def get_pending_qty(work_order,purpose):
	if purpose == "Material Transfer for Manufacture By FG":
		qty = frappe.db.sql("""SELECT (qty-material_transferred_for_manufacturing) as qty from `tabWork Order` where name = '{0}'""".format(work_order),as_dict=1)
		doc = frappe.get_doc("Work Order",work_order)
		ohs = get_current_stock()
		qty_will_be_produced_list = []
		for item in doc.required_items:
			if item.item_code in ohs:
				if item.required_qty <= ohs.get(item.item_code):
					percent_stock = 100
					qty_will_be_produced = item.required_qty
					qty_will_be_produced_list.append(qty_will_be_produced)
				elif item.required_qty > ohs.get(item.item_code) and ohs.get(item.item_code) > 0: 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = (percent_stock/100*item.required_qty)
					qty_will_be_produced_list.append(qty_will_be_produced)
				else : 
					percent_stock = (ohs.get(item.item_code)/item.required_qty*100)
					qty_will_be_produced = 0
					qty_will_be_produced_list.append(qty_will_be_produced)
		
		qty[0]['qty_will_be_produced'] =min(qty_will_be_produced_list)
		return qty
	else :
		qty = frappe.db.sql("""SELECT (qty-produced_qty) as qty from `tabWork Order` where name = '{0}'""".format(work_order))
		return qty

def get_sub_assembly_items(bom_no, bom_data, to_produce_qty, indent=0):
	data = get_children(parent = bom_no)
	for d in data:
		if d.expandable:
			parent_item_code = frappe.get_cached_value("BOM", bom_no, "item")
			stock_qty = (d.stock_qty / d.parent_bom_qty) * flt(to_produce_qty)
			bom_data.append(frappe._dict({
				'parent_item_code': parent_item_code,
				'description': d.description,
				'production_item': d.item_code,
				'item_name': d.item_name,
				'stock_uom': d.stock_uom,
				'uom': d.stock_uom,
				'bom_no': d.value,
				'is_sub_contracted_item': d.is_sub_contracted_item,
				'bom_level': indent,
				'indent': indent,
				'stock_qty': stock_qty
			}))

			if d.value:
				get_sub_assembly_items(d.value, bom_data, stock_qty, indent=indent+1)



@frappe.whitelist()
def create_stock_entry_for_po(purchase_order, item, row_name, name):
	po_raw_data = frappe.db.sql("""SELECT main_item, item_code, warehouse, purchase_order, picked_qty, batch_no, serial_no from `tabPurchase Order Pick List Item` where purchase_order='{0}' and main_item='{1}' and parent='{2}' """.format(purchase_order, item, name), as_dict=1)

	stock_entry = frappe.db.sql("""SELECT a.purchase_order, b.subcontracted_item from `tabStock Entry` a left join `tabStock Entry Detail` b on a.name=b.parent where a.purchase_order='{0}' and b.subcontracted_item='{1}'""".format(purchase_order, item), as_dict=True)
	t_warehouse = frappe.db.get_value("Purchase Order", purchase_order, "supplier_warehouse")
	wip_warehouse = frappe.db.get_single_value("Manufacturing Settings",'default_wip_warehouse')
	if len(stock_entry)==0:
		subcontracting_order = frappe.db.get_value("Subcontracting Order",{'purchase_order':purchase_order},'name')
		doc = frappe.new_doc("Stock Entry")
		doc.stock_entry_type = "Send to Subcontractor"
		# doc.purchase_order = purchase_order
		doc.subcontracting_order = subcontracting_order
		for row in po_raw_data:
			if t_warehouse != row.get("warehouse"): 
				doc.append("items", {
					"subcontracted_item": row.get("main_item"),
					"item_code":row.get("item_code"),
					"qty": row.get("picked_qty"),
					"s_warehouse": row.get("warehouse"),
					"t_warehouse": wip_warehouse,
					"basic_rate": frappe.db.get_value("Purchase Order Item Supplied", {"parent":purchase_order, "main_item_code":row.get("main_item"), "rm_item_code":row.get("item_code")}, "rate"),
					"batch_no": row.get("batch_no"),
					"serial_no": row.get("serial_no")
				})	
		doc.save()
		# doc.submit()
		status = "Submitted" if doc.docstatus==1 else "Draft"
		frappe.db.set_value("Pick List Purchase Order Table", {"name":row_name}, {"stock_entry_status":status, "stock_entry":doc.name})
		frappe.db.commit()

		frappe.msgprint("Stock Entries created {0}".format(doc.name))
		return {"name":doc.name, 'status':status}
	else:
		frappe.msgprint("Stock Entries already created for Purchase Order {0}".format(purchase_order))


@frappe.whitelist()
def create_dn_for_so(sales_order, item, row_name, name):
	so_data = frappe.db.sql("""SELECT item_code, warehouse, sales_order, picked_qty, batch_no, serial_no from `tabSales Order Pick List Item` where sales_order='{0}' and item_code='{1}' and parent='{2}' """.format(sales_order, item, name), as_dict=1)
	customer = frappe.db.get_value("Sales Order", sales_order, "customer")

	dn_entry = frappe.db.sql("""SELECT name From `tabDelivery Note Item` where against_sales_order='{0}' and item_code='{1}'""".format(sales_order, item), as_dict=1)

	if len(dn_entry)==0:
		so_doc = frappe.get_doc("Sales Order", sales_order)
		doc = frappe.new_doc("Delivery Note")
		doc.naming_series = "MAT-DN-.YYYY.-"
		doc.customer = so_doc.customer
		doc.woocommerce_order_id = so_doc.woocommerce_order_id
		doc.woocommerce_order_id = so_doc.woocommerce_order_id
		doc.currency = so_doc.currency
		doc.selling_price_list = so_doc.selling_price_list
		doc.selling_price_list = so_doc.selling_price_list
		doc.taxes_and_charges = "US ST 6% - RI"
		doc.tc_name = so_doc.tc_name
		doc.terms = so_doc.terms
		for row in so_doc.items:
			if row.item_code == item:
				for itm in so_data:
					doc.append("items", {
						"item_code": row.item_code,
						"item_name": row.item_name,
						"description": row.description,
						"uom": row.uom,
						"stock_uom": row.stock_uom,
						"conversion_factor": row.conversion_factor,
						"qty": itm.get("picked_qty"),
						"rate": row.rate,
						"warehouse": itm.get("warehouse"),
						"batch_no": itm.get("batch_no"),
						"against_sales_order": sales_order,
						"serial_no": itm.get("serial_no")
					})

		for row in so_doc.taxes:
			doc.append("taxes", {
				"charge_type": row.charge_type,
				"account_head": row.account_head,	
				"description": row.account_head,
				"rate": row.rate
			})
		doc.save()
		# doc.submit()

		if doc.name:
			status = "To Bill" if doc.docstatus==1 else "Draft" 
			frappe.msgprint("Delivery Note created {0}".format(doc.name))
			frappe.db.set_value("Pick List Sales Order Table", {"name":row_name}, {"delivery_note":doc.name, "delivery_note_status":status})
			frappe.db.commit()
			return {"name":doc.name, 'status':status}
	else:
		frappe.msgprint("Delivery Note already created for Sales Order {0}".format(sales_order))

@frappe.whitelist()
def create_stock_entries_in_bulk(doc):
	frappe.enqueue(
				"instrument.instrument.doctype.consolidated_pick_list.consolidated_pick_list.create_stock_entri",
				doc = doc,
				queue='long',
				timeout=1500)
@frappe.whitelist()
def create_stock_entri(doc):
	try:
		doc = json.loads(doc)
		work_orders = []
		work_orderss = frappe.db.sql("""SELECT wo.work_order,wo.bom_level,wo.name from `tabPick Orders` wo where wo.parent = '{0}' order by wo.bom_level,wo.name""".format(doc.get('name')),as_dict=1)
		work_order_list = [wo.work_order for wo in work_orderss]
		work_order_list = "', '".join(work_order_list)
		if doc.get('purpose')in ['Material Transfer for Manufacture By FG','Material Transfer for Manufacture By Work Order']:
			if doc.get('work_orders'):
				count = 0
				for row in doc.get('work_orders'):
					se_list = []
					stock_entry_list = frappe.db.sql("""SELECT se.name from `tabStock Entry` se where se.work_order = '{0}' and se.consolidated_pick_list = '{1}' and se.docstatus != '2' and se.purpose = 'Material Transfer for Manufacture'""".format(row.get('work_order'),doc.get('name')),as_dict=1)
					if len(stock_entry_list)!=0:
						continue
					else:
						se = create_stock_entry(row.get('work_order'),doc.get('name'),row.get('name'))
						if se:
							work_orders.append(se)
							se_list.append(se)
							count+=1
							se_list = [get_link_to_form("Stock Entry", p) for p in se_list]
							msgprint(_("{0} created").format(comma_and(se_list)))
							# if count == 15:
							# 	break
				if work_orders:
					work_orders = [get_link_to_form("Stock Entry", p) for p in work_orders]
					msgprint(_("{0} created").format(comma_and(work_orders)))
		elif doc.get('purpose') in ['Manufacture By FG','Manufacture By Work Order']:
			final_work_orders = []
			if doc.get('work_orders'):
				count = 0	
				if work_orderss:
					final_wo_dict = dict()
					for row in work_orderss:
						if row.get("bom_level") in final_wo_dict:
							updated_data = final_wo_dict.get(row.get("bom_level"))
							updated_data.append({'work_order':row.get('work_order'),'row_name':row.get('name')})
						else:
							final_wo_dict[row.get("bom_level")] = [{'work_order':row.get('work_order'),'row_name':row.get('name')}]
				if final_wo_dict:
					for row in final_wo_dict:
						for col in final_wo_dict.get(row):
							stock_entry_list = frappe.db.sql("""SELECT se.name from `tabStock Entry` se where se.work_order = '{0}' and se.consolidated_pick_list = '{1}' and se.docstatus != '2' and se.purpose = 'Manufacture'""".format(col.get('work_order'),doc.get('name')),as_dict=1)
							wo_doc =frappe.get_doc("Work Order",col.get('work_order'))
							if len(stock_entry_list)!=0:
								continue
							else:
								check_dependent_wo = frappe.db.sql("""SELECT wo.name from `tabWork Order` wo where wo.production_planning_with_lead_time = '{0}' and wo.so_reference = '{1}' and wo.produced_qty = 0 and wo.bom_level < {2} and wo.parent_item_code = '{3}' and wo.status != 'Completed' and wo.name in ('{4}')""".format(wo_doc.production_planning_with_lead_time,wo_doc.so_reference,wo_doc.bom_level,wo_doc.parent_item_code,work_order_list),as_dict=1,debug=1)
								check_dependent_wo = [i.name for i in check_dependent_wo]
								if check_dependent_wo and len(check_dependent_wo)>=1:
									frappe.throw("Please Complete Dependent Work Order Before Proceeding {0}".format(check_dependent_wo))
									break
								else:
									se = create_stock_entry(col.get('work_order'),doc.get('name'),col.get('row_name'))
								if se:
									final_work_orders.append(se)
									count+=1
									final_work_orders = [get_link_to_form("Stock Entry", p) for p in final_work_orders]
									msgprint(_("{0} created").format(comma_and(final_work_orders)))
									# if count == 15:
									# 	break
						
						if final_work_orders:
							final_work_orderss = "', '".join(final_work_orders)
							check_status = frappe.db.sql("""SELECT se.name from `tabStock Entry` se where se.docstatus = 0 and se.consolidated_pick_list = '{0}' and se.purpose = 'Manufacture' and se.name in ('{1}')""".format(doc.get('name'),final_work_orderss),as_dict=1)
							if check_status and len(check_status)>=1:
								msgprint(_("{0} Please Submit Stock Entries Before Proceed").format(comma_and(final_work_orders)))
	except Exception as e:
		traceback = frappe.get_traceback()
		frappe.log_error(
			title=_("Error while creating stock Entry"),
			message=traceback,
		)
		raise e					
@frappe.whitelist()
@frappe.validate_and_sanitize_search_inputs
def get_se(doctype, txt, searchfield, start, page_len, filters):
	if filters.get("work_order"):
		return frappe.db.sql(""" SELECT se.name,se.purpose FROM `tabStock Entry` se where se.work_order = '{0}'""".format(filters.get('work_order')),debug=1)